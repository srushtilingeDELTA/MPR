"""Parse New GSE MPR Workings.xlsx → GIR tab into values for the native PPT slide.

Fills the template chart/tables directly (no screenshots). The Workings GIR
dashboard mirrors the PowerPoint layout: summary, Injury Breakdown, Recordable,
and Metric.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from scorecard_screenshots import _find_sheet_name, resolve_sheet_name

logger = logging.getLogger(__name__)

MONTH_TOKENS = [
    "jan",
    "feb",
    "mar",
    "apr",
    "may",
    "jun",
    "jul",
    "aug",
    "sep",
    "oct",
    "nov",
    "dec",
]


@dataclass
class GirDashboard:
    """Values read from Workings!GIR for the template GIR slide."""

    month_label: str = ""
    actual: float | None = None
    plan: float | None = None
    yo1y: float | None = None
    yo2y: float | None = None
    injury_rows: list[list[str]] = field(default_factory=list)
    gir_monthly: list[float | None] = field(default_factory=lambda: [None] * 12)
    injury_monthly: list[float | None] = field(default_factory=lambda: [None] * 12)
    gir_ytd: float | None = None
    injury_ytd: float | None = None
    ytd_goal_var: float | None = None
    mtd_actual: float | None = None
    mtd_goal: float | None = None
    mtd_var: float | None = None
    ytd_actual: float | None = None
    ytd_goal: float | None = None
    ytd_var: float | None = None
    score_mtd: float | None = None
    score_ytd: float | None = None
    p1y_monthly: list[float | None] = field(default_factory=lambda: [None] * 12)
    p1y_ytd: float | None = None

    def has_core_values(self) -> bool:
        return any(
            v is not None
            for v in (
                self.actual,
                self.mtd_actual,
                self.gir_ytd,
                self.ytd_actual,
            )
        ) or any(v is not None for v in self.gir_monthly)


def _cell(ws: Worksheet, row: int, col: int):
    return ws.cell(row, col).value


def _cell_str(ws: Worksheet, row: int, col: int) -> str:
    val = _cell(ws, row, col)
    if val is None:
        return ""
    return str(val).strip()


def _as_float(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    neg = False
    if text.startswith("(") and text.endswith(")"):
        neg = True
        text = text[1:-1]
    text = text.replace("%", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if neg else number


def _fmt_number(value: float | None, decimals: int = 2) -> str:
    if value is None:
        return ""
    if decimals == 0:
        return f"{round(value):.0f}"
    return f"{value:.{decimals}f}"


def _fmt_var(value: float | None, decimals: int = 2) -> str:
    if value is None:
        return ""
    if value < 0:
        return f"({abs(value):.{decimals}f})"
    return _fmt_number(value, decimals)


def _find_token(
    ws: Worksheet,
    tokens: list[str],
    *,
    max_row: int = 80,
    max_col: int = 30,
) -> tuple[int, int] | None:
    needles = [t.casefold() for t in tokens]
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            text = _cell_str(ws, row, col).casefold()
            if not text:
                continue
            for needle in needles:
                if needle == text or needle in text:
                    return row, col
    return None


def _dashboard_end_col(ws: Worksheet) -> int:
    markers = ("yr_nb", "all", "no 115", "no115", "no. 115", "mo_nb", "kpi")
    for row in range(1, 40):
        for col in range(8, min(50, int(ws.max_column or 50)) + 1):
            text = _cell_str(ws, row, col).casefold()
            if text in markers or text.startswith("no 115"):
                return max(7, col - 1)
    return min(16, int(ws.max_column or 16))


def _row_values(ws: Worksheet, row: int, start_col: int, end_col: int) -> list:
    return [_cell(ws, row, col) for col in range(start_col, end_col + 1)]


def _parse_summary(ws: Worksheet, max_col: int, dash: GirDashboard) -> None:
    hit = _find_token(ws, ["yo1y", "yo2y", "yoy actuals", "actual:"], max_col=max_col)
    if not hit:
        return
    row0, col0 = hit
    # Scan a window around the YoY/Actual labels.
    for row in range(max(1, row0 - 6), min(int(ws.max_row or 40), row0 + 6) + 1):
        for col in range(max(1, col0 - 2), min(max_col, col0 + 3) + 1):
            label = _cell_str(ws, row, col).casefold().rstrip(":")
            right = _cell(ws, row, col + 1)
            below = _cell(ws, row + 1, col) if row < (ws.max_row or 1) else None
            value = right if _as_float(right) is not None or (isinstance(right, str) and right.strip()) else below
            if label in ("actual", "actual:"):
                dash.actual = _as_float(value) if _as_float(value) is not None else dash.actual
            elif label in ("plan", "plan:", "goal", "goal:"):
                dash.plan = _as_float(value) if _as_float(value) is not None else dash.plan
            elif label.startswith("yo1y"):
                dash.yo1y = _as_float(value) if _as_float(value) is not None else dash.yo1y
            elif label.startswith("yo2y"):
                dash.yo2y = _as_float(value) if _as_float(value) is not None else dash.yo2y
            elif re.search(r"[a-z]{3}'?\d{2}", label) or re.match(r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)", label):
                if not dash.month_label and "'" in _cell_str(ws, row, col):
                    dash.month_label = _cell_str(ws, row, col)


def _parse_injury(ws: Worksheet, max_col: int, dash: GirDashboard) -> None:
    hdr = _find_token(ws, ["injury breakdown"], max_col=max_col)
    year_hdr = _find_token(ws, ["nonrec", "non-rec", "dart"], max_col=max_col)
    if not hdr and not year_hdr:
        return
    title_row = hdr[0] if hdr else max(1, (year_hdr[0] if year_hdr else 2) - 1)
    # Header row with Year/Total/Rec...
    header_row = None
    start_col = 1
    for row in range(title_row, min(title_row + 4, int(ws.max_row or 40) + 1)):
        for col in range(1, max_col + 1):
            text = _cell_str(ws, row, col).casefold()
            if text == "year" or (text == "total" and _cell_str(ws, row, col + 1).casefold() in ("rec", "recordable")):
                header_row = row
                start_col = col if text == "year" else max(1, col - 1)
                break
        if header_row:
            break
    if header_row is None:
        header_row = title_row + 1
        start_col = hdr[1] if hdr else 1

    end_col = start_col + 4
    rows: list[list[str]] = []
    for row in range(header_row + 1, min(header_row + 12, int(ws.max_row or 80) + 1)):
        vals = _row_values(ws, row, start_col, end_col)
        if not any(v is not None and str(v).strip() != "" for v in vals):
            if rows:
                break
            continue
        formatted = []
        for idx, val in enumerate(vals):
            if val is None or str(val).strip() == "":
                formatted.append("")
            elif idx == 0:
                formatted.append(str(val).strip())
            else:
                num = _as_float(val)
                formatted.append(_fmt_number(num, 0) if num is not None else str(val).strip())
        rows.append(formatted)
        if formatted[0].casefold() == "total":
            break
    dash.injury_rows = rows


def _month_col_map(ws: Worksheet, header_row: int, start_col: int, end_col: int) -> dict[int, int]:
    """Map month number 1-12, and special keys 13=YTD, 14=YTD Goal Var → excel col."""
    mapping: dict[int, int] = {}
    for col in range(start_col, end_col + 1):
        text = _cell_str(ws, header_row, col).casefold().replace(".", "")
        if not text:
            continue
        for idx, token in enumerate(MONTH_TOKENS, start=1):
            if text.startswith(token):
                mapping[idx] = col
                break
        if "ytd" in text and "goal" in text:
            mapping[14] = col
        elif text == "ytd" or text.startswith("ytd"):
            mapping[13] = col
    return mapping


def _parse_recordable(ws: Worksheet, max_col: int, dash: GirDashboard) -> None:
    hit = _find_token(ws, ["recordable"], max_col=max_col)
    if not hit:
        return
    header_row, start_col = hit
    # Month headers are usually on the same row as Recordable.
    col_map = _month_col_map(ws, header_row, start_col, max_col)
    if len(col_map) < 3:
        # Try next row.
        col_map = _month_col_map(ws, header_row + 1, 1, max_col)
        if len(col_map) >= 3:
            header_row = header_row + 1

    gir_row = injury_row = None
    for row in range(header_row + 1, min(header_row + 8, int(ws.max_row or 80) + 1)):
        label = _cell_str(ws, row, 1).casefold()
        if not label:
            label = _cell_str(ws, row, start_col).casefold()
        # Stop if we've left the Recordable block.
        if label in ("metric", "month to date") or "injury breakdown" in label:
            break
        if "system gir" in label:
            gir_row = row
        elif injury_row is None and "injury count" in label:
            injury_row = row
        elif injury_row is None and label.startswith("injury"):
            injury_row = row
        # Do not treat the Metric table's bare "GIR" row as System GIR.
    def _read_row(row: int | None, as_int: bool = False) -> tuple[list[float | None], float | None, float | None]:
        monthly = [None] * 12
        ytd = var = None
        if row is None:
            return monthly, ytd, var
        for month, col in col_map.items():
            val = _as_float(_cell(ws, row, col))
            if month <= 12:
                monthly[month - 1] = val
            elif month == 13:
                ytd = val
            elif month == 14:
                var = val
        if as_int:
            monthly = [None if v is None else float(round(v)) for v in monthly]
            if ytd is not None:
                ytd = float(round(ytd))
        return monthly, ytd, var

    dash.gir_monthly, dash.gir_ytd, dash.ytd_goal_var = _read_row(gir_row)
    dash.injury_monthly, dash.injury_ytd, _ = _read_row(injury_row, as_int=True)


def _parse_metric(ws: Worksheet, max_col: int, dash: GirDashboard) -> None:
    hit = _find_token(ws, ["month to date", "metric"], max_col=max_col)
    if not hit:
        return
    row0, _ = hit
    # Find the GIR data row near Metric / Month to Date.
    data_row = None
    header_row = None
    for row in range(max(1, row0 - 2), min(row0 + 6, int(ws.max_row or 80) + 1)):
        for col in range(1, min(4, max_col + 1)):
            text = _cell_str(ws, row, col).casefold()
            if text == "gir" or text.startswith("gir"):
                data_row = row
            if "actual" in text and header_row is None:
                header_row = row
    if data_row is None:
        return

    # Expected layout: Actual Goal +/- | Actual Goal +/- | Score MTD YTD
    # Find columns by scanning header rows above the data row.
    headers: dict[int, str] = {}
    for row in range(max(1, data_row - 3), data_row):
        for col in range(1, max_col + 1):
            text = _cell_str(ws, row, col).casefold()
            if text:
                headers[col] = text

    # Fallback positional mapping used by the PPT/Excel mirror layout.
    values = [_as_float(_cell(ws, data_row, col)) for col in range(2, 10)]
    while len(values) < 8:
        values.append(None)
    (
        dash.mtd_actual,
        dash.mtd_goal,
        dash.mtd_var,
        dash.ytd_actual,
        dash.ytd_goal,
        dash.ytd_var,
        dash.score_mtd,
        dash.score_ytd,
    ) = values[:8]

    # Prefer labeled columns when available.
    actual_cols = [c for c, t in headers.items() if t == "actual"]
    goal_cols = [c for c, t in headers.items() if t == "goal"]
    var_cols = [c for c, t in headers.items() if "+/-" in t or "vs goal" in t]
    score_cols = [c for c, t in headers.items() if t in ("mtd", "ytd") and c > 6]
    if len(actual_cols) >= 1:
        dash.mtd_actual = _as_float(_cell(ws, data_row, actual_cols[0]))
    if len(actual_cols) >= 2:
        dash.ytd_actual = _as_float(_cell(ws, data_row, actual_cols[1]))
    if len(goal_cols) >= 1:
        dash.mtd_goal = _as_float(_cell(ws, data_row, goal_cols[0]))
    if len(goal_cols) >= 2:
        dash.ytd_goal = _as_float(_cell(ws, data_row, goal_cols[1]))
    if len(var_cols) >= 1:
        dash.mtd_var = _as_float(_cell(ws, data_row, var_cols[0]))
    if len(var_cols) >= 2:
        dash.ytd_var = _as_float(_cell(ws, data_row, var_cols[1]))


def _parse_p1y_from_nearby(ws: Worksheet, max_col: int, dash: GirDashboard) -> None:
    """Optional prior-year monthly row on the dashboard."""
    hit = _find_token(ws, ["p1y", "prior year", "py actual", "prior yr"], max_col=max_col)
    if not hit:
        return
    row, start_col = hit
    # If this is a chart series label only, months may be on a header above.
    header_row = row - 1 if row > 1 else row
    col_map = _month_col_map(ws, header_row, 1, max_col)
    if len(col_map) < 3:
        col_map = _month_col_map(ws, row, 1, max_col)
    for month, col in col_map.items():
        val = _as_float(_cell(ws, row, col))
        if month <= 12:
            dash.p1y_monthly[month - 1] = val
        elif month == 13:
            dash.p1y_ytd = val


def parse_gir_dashboard(workbook_bytes: bytes, sheet_name: str) -> GirDashboard:
    """Parse the formatted GIR dashboard (left side; excludes ALL/NO 115 dumps)."""
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    try:
        match = _find_sheet_name(list(wb.sheetnames), sheet_name) or sheet_name
        if match not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
        ws = wb[match]
        max_col = _dashboard_end_col(ws)
        dash = GirDashboard()
        _parse_summary(ws, max_col, dash)
        _parse_injury(ws, max_col, dash)
        _parse_recordable(ws, max_col, dash)
        _parse_metric(ws, max_col, dash)
        _parse_p1y_from_nearby(ws, max_col, dash)

        # Cross-fill from related sections when one block is sparse.
        if dash.actual is None:
            dash.actual = dash.mtd_actual
        if dash.plan is None:
            dash.plan = dash.mtd_goal or dash.ytd_goal
        if dash.mtd_actual is None:
            dash.mtd_actual = dash.actual
        if dash.mtd_goal is None:
            dash.mtd_goal = dash.plan
        if dash.ytd_actual is None:
            dash.ytd_actual = dash.gir_ytd
        if dash.ytd_goal is None:
            dash.ytd_goal = dash.plan
        if dash.gir_ytd is None:
            dash.gir_ytd = dash.ytd_actual
        if dash.mtd_var is None and dash.mtd_actual is not None and dash.mtd_goal is not None:
            dash.mtd_var = dash.mtd_actual - dash.mtd_goal
        if dash.ytd_var is None and dash.ytd_actual is not None and dash.ytd_goal is not None:
            dash.ytd_var = dash.ytd_actual - dash.ytd_goal
        if dash.ytd_goal_var is None:
            dash.ytd_goal_var = dash.ytd_var

        logger.info(
            "Parsed Workings GIR: actual=%s plan=%s ytd=%s injury_ytd=%s injury_rows=%s",
            dash.actual,
            dash.plan,
            dash.gir_ytd,
            dash.injury_ytd,
            len(dash.injury_rows),
        )
        return dash
    finally:
        wb.close()


def load_gir_dashboard_from_data(data, element: dict | None = None) -> GirDashboard:
    """Resolve Workings!GIR and parse the dashboard."""
    element = element or {}
    workbook = element.get("workbook", "workings")
    workbook_bytes = data.store.workbook_bytes(workbook)
    available = []
    try:
        available = list(data.sheet_names(workbook))
    except Exception:
        available = []
    sheet_name = resolve_sheet_name(
        workbook_bytes,
        sheet=element.get("sheet", "GIR"),
        sheet_index=element.get("sheet_index"),
        sheet_match=element.get("sheet_match", ["GIR"]),
        sheet_match_index=int(element.get("sheet_match_index", 0) or 0),
        available=available or None,
    )
    return parse_gir_dashboard(workbook_bytes, sheet_name)


def apply_gir_workings_native(slide, data, element: dict) -> bool:
    """Fill native GIR chart + tables from Workings!GIR (no screenshots)."""
    from ppt_format import set_cell_text_preserve

    try:
        dash = load_gir_dashboard_from_data(data, element)
    except Exception as exc:
        logger.warning("Workings GIR parse failed: %s", exc)
        return False

    if not dash.has_core_values():
        logger.warning("Workings GIR dashboard had no usable values")
        return False

    month_label = dash.month_label or data.report_month_short()
    through_month = int(getattr(data, "month", 12) or 12)

    filled_tables = 0
    for shape in slide.shapes:
        if not getattr(shape, "has_table", False):
            continue
        table = shape.table
        label = (table.cell(0, 0).text or "").strip().casefold()

        if "injury breakdown" in label and dash.injury_rows:
            start_row = 2 if len(table.rows) > 2 else 1
            for r_idx, row in enumerate(dash.injury_rows):
                ppt_row = start_row + r_idx
                if ppt_row >= len(table.rows):
                    break
                for c_idx in range(min(len(table.columns), len(row))):
                    set_cell_text_preserve(table.cell(ppt_row, c_idx), row[c_idx])
            filled_tables += 1
            continue

        if label == "metric" or (
            len(table.rows) > 2 and (table.cell(2, 0).text or "").strip().upper() == "GIR"
        ):
            if len(table.rows) > 2 and len(table.columns) >= 7:
                set_cell_text_preserve(table.cell(2, 1), _fmt_number(dash.mtd_actual))
                set_cell_text_preserve(table.cell(2, 2), _fmt_number(dash.mtd_goal))
                set_cell_text_preserve(
                    table.cell(2, 3),
                    _fmt_var(dash.mtd_var)
                    if dash.mtd_var is not None
                    else _fmt_var(
                        None
                        if dash.mtd_actual is None or dash.mtd_goal is None
                        else dash.mtd_actual - dash.mtd_goal
                    ),
                )
                set_cell_text_preserve(table.cell(2, 4), _fmt_number(dash.ytd_actual))
                set_cell_text_preserve(table.cell(2, 5), _fmt_number(dash.ytd_goal))
                set_cell_text_preserve(
                    table.cell(2, 6),
                    _fmt_var(dash.ytd_var)
                    if dash.ytd_var is not None
                    else _fmt_var(
                        None
                        if dash.ytd_actual is None or dash.ytd_goal is None
                        else dash.ytd_actual - dash.ytd_goal
                    ),
                )
                if len(table.columns) > 7 and dash.score_mtd is not None:
                    set_cell_text_preserve(table.cell(2, 7), _fmt_number(dash.score_mtd))
                if len(table.columns) > 8 and dash.score_ytd is not None:
                    set_cell_text_preserve(table.cell(2, 8), _fmt_number(dash.score_ytd))
                filled_tables += 1
            continue

        # May / YoY summary (6x2-ish).
        first_body = (table.cell(1, 0).text or "").strip().casefold() if len(table.rows) > 1 else ""
        if "actual" in first_body or month_label.casefold() in label or "yoy" in "".join(
            (table.cell(r, 0).text or "").casefold() for r in range(len(table.rows))
        ):
            set_cell_text_preserve(table.cell(0, 0), month_label)
            for row_idx in range(len(table.rows)):
                left = (table.cell(row_idx, 0).text or "").strip().casefold()
                if left.startswith("actual"):
                    set_cell_text_preserve(table.cell(row_idx, 1), _fmt_number(dash.actual))
                elif left.startswith("plan") or left.startswith("goal"):
                    set_cell_text_preserve(table.cell(row_idx, 1), _fmt_number(dash.plan))
                elif left.startswith("yo1y"):
                    set_cell_text_preserve(table.cell(row_idx, 1), _fmt_number(dash.yo1y))
                elif left.startswith("yo2y"):
                    set_cell_text_preserve(table.cell(row_idx, 1), _fmt_number(dash.yo2y))
            filled_tables += 1
            continue

        if label == "recordable" or "system gir" in (
            table.cell(1, 0).text or ""
        ).casefold():
            for col_idx, month_num in enumerate(range(1, 13), start=1):
                if month_num <= through_month:
                    set_cell_text_preserve(
                        table.cell(1, col_idx), _fmt_number(dash.gir_monthly[month_num - 1])
                    )
                else:
                    set_cell_text_preserve(table.cell(1, col_idx), "")
            if len(table.columns) > 13:
                set_cell_text_preserve(table.cell(1, 13), _fmt_number(dash.gir_ytd))
            if len(table.columns) > 14:
                set_cell_text_preserve(table.cell(1, 14), _fmt_var(dash.ytd_goal_var))
            if len(table.rows) > 2:
                for col_idx, month_num in enumerate(range(1, 13), start=1):
                    if month_num <= through_month:
                        set_cell_text_preserve(
                            table.cell(2, col_idx),
                            _fmt_number(dash.injury_monthly[month_num - 1], decimals=0),
                        )
                    else:
                        set_cell_text_preserve(table.cell(2, col_idx), "")
                if len(table.columns) > 13:
                    set_cell_text_preserve(
                        table.cell(2, 13), _fmt_number(dash.injury_ytd, decimals=0)
                    )
                if len(table.columns) > 14:
                    set_cell_text_preserve(table.cell(2, 14), "")
            filled_tables += 1

    # Chart: Current Year + Goal from Workings; preserve/replace P1Y when available.
    from ppt_builder import update_gir_chart_from_series

    chart_updated = update_gir_chart_from_series(
        slide,
        monthly=dash.gir_monthly,
        goal=dash.plan or dash.mtd_goal or dash.ytd_goal,
        ytd=dash.gir_ytd or dash.ytd_actual,
        prior_year=dash.p1y_monthly if any(v is not None for v in dash.p1y_monthly) else None,
        prior_ytd=dash.p1y_ytd,
    )

    if bool(element.get("clear_narrative", True)):
        from gir_panels import clear_leading_action_narrative

        n = clear_leading_action_narrative(slide)
        print(f">>> GIR Leading Issues / Action Plan cleared ({n} text box(es))")

    print(
        f"\n>>> GIR slide filled from Workings!GIR "
        f"(tables={filled_tables}, chart={'yes' if chart_updated else 'no'}, "
        f"actual={_fmt_number(dash.actual)}, plan={_fmt_number(dash.plan)}, "
        f"ytd={_fmt_number(dash.gir_ytd)})\n"
    )
    return filled_tables > 0 or chart_updated
