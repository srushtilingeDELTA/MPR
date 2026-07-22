"""North Scorecard Summary (PPT 14) from GSE MPR Visualizations.xlsx.

Matches the template layout:
  - Left: two legend tables (score bands + KPI better/worse)
  - Center: GSE MPR category summary table on top
  - Center: KPI metrics table underneath (Global Injury Rate … Lead Input)

Columns are entity stations: NORTH, BOS (M), DTW (M), MSP (M), SLC (M), NY (M).
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image
from pptx.enum.shapes import MSO_SHAPE_TYPE

from scorecard_screenshots import (
    _fill_rgb,
    _find_sheet_name,
    _range_address,
    _remove_slide_pictures,
    _used_bounds,
    _validate_capture,
    capture_range_png,
    place_picture_on_slide,
    resolve_sheet_name,
)

logger = logging.getLogger(__name__)

# Measured content band from the live North Scorecard Summary template (EMU).
# Two tables stacked in the center; two legends stacked on the left.
_CONTENT_LEFT = 2_208_251
_CONTENT_TOP = 1_009_877
_CONTENT_WIDTH = 7_775_497
_CONTENT_HEIGHT = 4_838_246
_LEGEND_LEFT = 314_628
_LEGEND_WIDTH = 1_750_000
_GAP = 90_000
# Summary is shorter (GSE MPR + 6 category rows); metrics holds 12 KPI rows.
_SUMMARY_HEIGHT = int(_CONTENT_HEIGHT * 0.36)
_METRICS_HEIGHT = _CONTENT_HEIGHT - _SUMMARY_HEIGHT - _GAP

# Compact legend slots (match template: small tables, not stretched half-slide).
_SCORE_LEGEND_HEIGHT = 1_050_000  # Legend + 3 score rows
_KPI_LEGEND_HEIGHT = 1_300_000  # Legend + 4 KPI rows
_LEGEND_STACK_GAP = 220_000

NORTH_SUMMARY_BOX = (_CONTENT_LEFT, _CONTENT_TOP, _CONTENT_WIDTH, _SUMMARY_HEIGHT)
NORTH_METRICS_BOX = (
    _CONTENT_LEFT,
    _CONTENT_TOP + _SUMMARY_HEIGHT + _GAP,
    _CONTENT_WIDTH,
    _METRICS_HEIGHT,
)
NORTH_LEGEND_BOXES = (
    (_LEGEND_LEFT, _CONTENT_TOP, _LEGEND_WIDTH, _SCORE_LEGEND_HEIGHT),
    (
        _LEGEND_LEFT,
        _CONTENT_TOP + _SCORE_LEGEND_HEIGHT + _LEGEND_STACK_GAP,
        _LEGEND_WIDTH,
        _KPI_LEGEND_HEIGHT,
    ),
)

# Keep old name for any callers expecting a single main box.
NORTH_MAIN_BOX = (
    _CONTENT_LEFT,
    _CONTENT_TOP,
    _CONTENT_WIDTH,
    _CONTENT_HEIGHT,
)

CATEGORY_ROWS = (
    "safety",
    "customer experience",
    "operations",
    "finance",
    "people",
    "total",
)

KPI_KEYS = (
    "global injury rate",
    "ea compliance",
    "asap",
    "isr%",
    "sev",
    "pmi",
    "pmi nme",
    "qc compliance",
    "budget $000s",
    "overtime",
    "total hours",
    "lead input",
)

KPI_ALIASES: dict[str, tuple[str, ...]] = {
    "global injury rate": ("global injury rate", "gir", "injury rate"),
    "ea compliance": ("ea compliance", "eac", "ea comp"),
    "asap": ("asap",),
    "isr%": ("isr%", "isr %", "isr"),
    "sev": ("sev", "severity"),
    "pmi": ("pmi", "pm (m)", "pm motorized"),
    "pmi nme": ("pmi nme", "pm (nme)", "pm nme", "non-motorized", "non motorized"),
    "qc compliance": ("qc compliance", "qc comp", "qcc"),
    "budget $000s": (
        "budget $000s",
        "budget $000",
        "budget ($000s)",
        "budget ($000)",
        "budget",
    ),
    "overtime": ("overtime", "ot", "o.t."),
    "total hours": ("total hours", "hours", "total hrs", "tot hours"),
    "lead input": ("lead input", "lead inputs", "leadership input"),
}

ENTITY_TOKENS = (
    "north",
    "bos",
    "dtw",
    "msp",
    "slc",
    "ny",
    "jfk",
    "lga",
    "ewr",
)

SCORE_LEGEND_ROWS = (
    "score above 4",
    "score between 3 and 4",
    "score below 3",
)

KPI_LEGEND_ROWS = (
    "kpi better than goal",
    "kpi worse than goal",
    "kpi not applicable",
    "goal pending for kpi",
)

# Accept Excel wording variants (with/without "KPI" prefix, shorter forms).
SCORE_LEGEND_ALIASES: dict[str, tuple[str, ...]] = {
    "score above 4": ("score above 4", "above 4", "score > 4", "score >= 4"),
    "score between 3 and 4": (
        "score between 3 and 4",
        "between 3 and 4",
        "score 3 to 4",
        "3 and 4",
    ),
    "score below 3": ("score below 3", "below 3", "score < 3"),
}

KPI_LEGEND_ALIASES: dict[str, tuple[str, ...]] = {
    "kpi better than goal": (
        "kpi better than goal",
        "better than goal",
        "better than plan",
    ),
    "kpi worse than goal": (
        "kpi worse than goal",
        "worse than goal",
        "worse than plan",
    ),
    "kpi not applicable": (
        "kpi not applicable",
        "not applicable",
        "n/a",
        "na",
    ),
    "goal pending for kpi": (
        "goal pending for kpi",
        "goal pending",
        "pending for kpi",
        "pending goal",
    ),
}

# Backward-compatible aliases used for classification / fuzzy match.
SCORE_LEGEND_TOKENS = SCORE_LEGEND_ROWS + (
    "score above",
    "score between",
    "score below",
)

KPI_LEGEND_TOKENS = KPI_LEGEND_ROWS + (
    "better than goal",
    "worse than goal",
    "not applicable",
    "goal pending",
)

LEGEND_HEADER_TOKENS = ("legend", "color key", "status key", "score key", "key:")

LEGEND_MIN_EXTRA_ROWS = 3
LEGEND_MIN_EXTRA_COLS = 2
LEGEND_MAX_ROWS = 10
LEGEND_MAX_COLS = 5


@dataclass
class RangeBlock:
    label: str
    start_row: int
    end_row: int
    start_col: int
    end_col: int

    @property
    def range_address(self) -> str:
        return _range_address(self.start_row, self.end_row, self.start_col, self.end_col)

    @property
    def row_count(self) -> int:
        return self.end_row - self.start_row + 1

    @property
    def col_count(self) -> int:
        return self.end_col - self.start_col + 1


@dataclass
class NorthSummaryLayout:
    summary: RangeBlock
    metrics: RangeBlock
    legends: list[RangeBlock]


def _cell_str(ws, row: int, col: int) -> str:
    val = ws.cell(row, col).value
    if val is None:
        return ""
    return str(val).strip()


def _norm(text: str) -> str:
    cleaned = (
        str(text or "")
        .replace("\xa0", " ")
        .replace("\u2007", " ")
        .replace("\u202f", " ")
        .replace("\r", " ")
        .replace("\n", " ")
        .replace("\t", " ")
    )
    cleaned = re.sub(r"[()]", " ", cleaned)
    return re.sub(r"\s+", " ", cleaned).casefold().strip()


def _cell_has_fill(ws, row: int, col: int) -> bool:
    try:
        rgb = _fill_rgb(ws.cell(row, col))
    except Exception:
        return False
    if rgb is None:
        return False
    return not (rgb[0] >= 245 and rgb[1] >= 245 and rgb[2] >= 245)


def _cell_interesting(ws, row: int, col: int) -> bool:
    return bool(_cell_str(ws, row, col)) or _cell_has_fill(ws, row, col)


def _match_category(text: str, token: str) -> bool:
    n = _norm(text)
    if not n:
        return False
    if token == "total":
        return n == "total" or (n.startswith("total ") and "hour" not in n)
    if token == "safety":
        return n == "safety" or n.startswith("safety")
    return token == n or token in n


def _match_kpi(text: str, key: str) -> bool:
    n = _norm(text)
    if not n:
        return False
    for alias in KPI_ALIASES.get(key, (key,)):
        a = _norm(alias)
        if not a:
            continue
        if len(a) <= 3:
            tokens = set(re.findall(r"[a-z0-9%]+", n))
            if a not in tokens and n != a:
                continue
        else:
            if not (n == a or a in n or a.replace(" ", "") in n.replace(" ", "")):
                continue
        if key == "pmi" and "nme" in n:
            continue
        if key == "lead input" and "leading" in n:
            continue
        if key == "isr%" and n in {"severity", "sev"}:
            continue
        if key == "overtime" and "total" in n and "overtime" not in n:
            tokens = set(re.findall(r"[a-z0-9%]+", n))
            if "ot" not in tokens:
                continue
        return True
    return False


def _is_entity_header(text: str) -> bool:
    n = _norm(text)
    if not n:
        return False
    tokens = set(re.findall(r"[a-z0-9]+", n))
    return any(tok in tokens or tok == n for tok in ENTITY_TOKENS)


def _is_legend_header(text: str) -> bool:
    n = _norm(text)
    return bool(n) and any(tok in n for tok in LEGEND_HEADER_TOKENS)


def _match_legend_row(text: str, token: str, *, aliases: dict[str, tuple[str, ...]] | None = None) -> bool:
    n = _norm(text)
    if not n:
        return False
    options = aliases.get(token, (token,)) if aliases else (token,)
    for opt in options:
        t = _norm(opt)
        if not t:
            continue
        if n == t or t in n:
            return True
        # Allow cell text that is the alias without a longer prefix (e.g. "Better than Goal").
        if len(t) >= 8 and n in t:
            return True
    return False


def _find_legend_row_cells(
    ws,
    tokens: tuple[str, ...],
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    *,
    aliases: dict[str, tuple[str, ...]] | None = None,
    exclude: tuple[int, int, int, int] | None = None,
) -> list[tuple[int, int, str]]:
    """Locate cells matching the known legend body labels (outside the main tables)."""
    hits: list[tuple[int, int, str]] = []
    used_rows: set[int] = set()

    def in_exclude(r: int, c: int) -> bool:
        if exclude is None:
            return False
        er0, er1, ec0, ec1 = exclude
        return er0 <= r <= er1 and ec0 <= c <= ec1

    for token in tokens:
        for row in range(min_row, max_row + 1):
            if row in used_rows:
                continue
            for col in range(min_col, max_col + 1):
                if in_exclude(row, col):
                    continue
                if _match_legend_row(_cell_str(ws, row, col), token, aliases=aliases):
                    hits.append((row, col, token))
                    used_rows.add(row)
                    break
            else:
                continue
            break
    return hits


def _is_legend_junk_neighbor(text: str) -> bool:
    """True for nearby Excel labels that must not be pulled into a legend crop."""
    n = _norm(text)
    if not n:
        return False
    tokens = set(re.findall(r"[a-z0-9%]+", n))
    junk = {
        "ot",
        "hours",
        "hc",
        "fte",
        "headcount",
        "budget",
        "north",
        "bos",
        "dtw",
        "msp",
        "slc",
        "ny",
        "jfk",
        "lga",
        "safety",
        "operations",
        "finance",
        "people",
        "total",
    }
    if tokens & junk:
        return True
    # Longer labels that are clearly not legend body text.
    for bad in (
        "overtime",
        "total hours",
        "lead input",
        "global injury",
        "customer experience",
        "gse mpr",
    ):
        if bad in n:
            return True
    return False


def _legend_block_from_rows(
    ws,
    *,
    kind: str,
    body_hits: list[tuple[int, int, str]],
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    exclude: tuple[int, int, int, int] | None = None,
) -> RangeBlock | None:
    """Build a tight legend capture: Legend header + swatch + known body rows only.

    Does not expand into neighboring Excel junk (OT / Hours / HC / scorecard cols).
    """
    if not body_hits:
        return None

    # Prefer the densest same-column cluster of body labels.
    by_col: dict[int, list[tuple[int, int, str]]] = {}
    for hit in body_hits:
        by_col.setdefault(hit[1], []).append(hit)
    text_col, cluster = max(by_col.items(), key=lambda kv: len(kv[1]))
    rows = [r for r, _c, _t in cluster]
    start_row = min(rows)
    end_row = max(rows)
    start_col = text_col
    end_col = text_col

    def in_exclude(r: int, c: int) -> bool:
        if exclude is None:
            return False
        er0, er1, ec0, ec1 = exclude
        return er0 <= r <= er1 and ec0 <= c <= ec1

    # Color-swatch column immediately left of the text (fills only / Legend header).
    if start_col > min_col:
        left = start_col - 1
        left_texts = [_cell_str(ws, r, left) for r in range(start_row, end_row + 1)]
        if any(_is_legend_junk_neighbor(t) for t in left_texts):
            pass
        elif any(
            (_cell_has_fill(ws, r, left) or not _cell_str(ws, r, left))
            and not in_exclude(r, left)
            for r in range(start_row, end_row + 1)
        ) or any(
            _is_legend_header(_cell_str(ws, r, left))
            for r in range(max(min_row, start_row - 2), start_row + 1)
        ):
            start_col = left

    # "Legend" header row directly above the body (do not skip into unrelated blocks).
    header_row = None
    for r in range(start_row - 1, max(min_row, start_row - 3) - 1, -1):
        header_hit = False
        for c in range(start_col, end_col + 1):
            if in_exclude(r, c):
                continue
            text = _cell_str(ws, r, c)
            if _is_legend_header(text):
                header_row = r
                header_hit = True
                break
            if text and _is_legend_junk_neighbor(text):
                header_hit = True  # stop upward scan
                break
        if header_hit:
            break
        # Blank / fill-only row — keep looking one more step.
    if header_row is not None:
        start_row = header_row

    # Include merged cells that only cover the legend footprint (not neighbors).
    try:
        for merged in ws.merged_cells.ranges:
            if merged.min_row < start_row or merged.max_row > end_row:
                continue
            if merged.min_col < start_col - 0 or merged.max_col > end_col + 1:
                # Allow header merge one col past text.
                if not (
                    merged.min_row == start_row
                    and merged.min_col >= start_col
                    and merged.max_col <= end_col + 1
                ):
                    continue
            # Reject merges that bring in junk text.
            junk = False
            for r in range(int(merged.min_row), int(merged.max_row) + 1):
                for c in range(int(merged.min_col), int(merged.max_col) + 1):
                    if _is_legend_junk_neighbor(_cell_str(ws, r, c)):
                        junk = True
                        break
                if junk:
                    break
            if junk:
                continue
            start_col = min(start_col, int(merged.min_col))
            end_col = max(end_col, int(merged.max_col))
    except Exception:
        pass

    # Hard cap: never more columns than swatch + label (+ optional wrap).
    if end_col - start_col > 2:
        end_col = start_col + 2

    return RangeBlock(
        label=f"legend_{kind}",
        start_row=start_row,
        end_row=end_row,
        start_col=start_col,
        end_col=end_col,
    )


def _legend_kind(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> str:
    """Classify a legend block as score-band vs KPI status."""
    blob = " ".join(
        _norm(_cell_str(ws, r, c))
        for r in range(start_row, end_row + 1)
        for c in range(start_col, end_col + 1)
    )
    score_hits = sum(1 for tok in SCORE_LEGEND_TOKENS if tok in blob)
    kpi_hits = sum(1 for tok in KPI_LEGEND_TOKENS if tok in blob)
    if score_hits > kpi_hits:
        return "score"
    if kpi_hits > score_hits:
        return "kpi"
    return "unknown"


def _find_row_labels(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> dict[str, tuple[int, int]]:
    found: dict[str, tuple[int, int]] = {}
    for row in range(min_row, max_row + 1):
        for col in range(min_col, min(max_col, min_col + 4) + 1):
            text = _cell_str(ws, row, col)
            if not text:
                continue
            if "gse mpr" not in found and "gse mpr" in _norm(text):
                found["gse mpr"] = (row, col)
            for token in CATEGORY_ROWS:
                if token not in found and _match_category(text, token):
                    found[token] = (row, col)
            for key in KPI_KEYS:
                if key not in found and _match_kpi(text, key):
                    found[key] = (row, col)
    return found


def _find_entity_header_row(
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    *,
    near_row: int | None = None,
) -> tuple[int, int, int] | None:
    """Return (header_row, first_entity_col, last_entity_col) near a label block."""
    best = None
    for row in range(min_row, max_row + 1):
        cols = [c for c in range(min_col, max_col + 1) if _is_entity_header(_cell_str(ws, row, c))]
        if len(cols) < 2:
            continue
        score = len(cols)
        if near_row is not None:
            score -= abs(row - near_row) * 0.01
        if best is None or score > best[0]:
            best = (score, row, min(cols), max(cols))
    if best is None:
        return None
    return best[1], best[2], best[3]


def _expand_table(
    ws,
    seed_rows: list[int],
    seed_cols: list[int],
    *,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    header_row: int | None = None,
    entity_end_col: int | None = None,
    stop_before_rows: set[int] | None = None,
) -> RangeBlock:
    start_row = min(seed_rows)
    end_row = max(seed_rows)
    start_col = min(seed_cols)
    end_col = max(seed_cols)

    if header_row is not None:
        start_row = min(start_row, header_row)
    if entity_end_col is not None:
        end_col = max(end_col, entity_end_col)

    # Include title row above header when present (e.g. GSE MPR banner).
    if start_row > min_row:
        above = start_row - 1
        if any(_cell_interesting(ws, above, c) for c in range(start_col, end_col + 1)):
            start_row = above

    stop_before_rows = stop_before_rows or set()
    empty = 0
    cursor = end_row
    while cursor < max_row and empty < 2:
        nxt = cursor + 1
        if nxt in stop_before_rows:
            break
        if any(_cell_interesting(ws, nxt, c) for c in range(start_col, end_col + 1)):
            end_row = nxt
            cursor = nxt
            empty = 0
        else:
            empty += 1
            cursor = nxt

    while end_col < max_col and (entity_end_col is None or end_col < entity_end_col):
        nxt = end_col + 1
        if entity_end_col is not None and nxt > entity_end_col:
            break
        if any(_cell_interesting(ws, r, nxt) for r in range(start_row, end_row + 1)):
            end_col = nxt
            continue
        break
    if entity_end_col is not None:
        end_col = max(end_col, entity_end_col)

    try:
        for merged in ws.merged_cells.ranges:
            if (
                merged.max_row >= start_row
                and merged.min_row <= end_row
                and merged.max_col >= start_col
                and merged.min_col <= end_col
            ):
                start_row = min(start_row, int(merged.min_row))
                end_row = max(end_row, int(merged.max_row))
                start_col = min(start_col, int(merged.min_col))
                end_col = max(end_col, int(merged.max_col))
    except Exception:
        pass

    return RangeBlock(
        label="table",
        start_row=start_row,
        end_row=end_row,
        start_col=start_col,
        end_col=end_col,
    )


def _expand_legend_table(
    ws,
    anchor_row: int,
    anchor_col: int,
    *,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    exclude: tuple[int, int, int, int] | None = None,
    other_legend_anchors: list[tuple[int, int]] | None = None,
) -> RangeBlock | None:
    def in_exclude(r: int, c: int) -> bool:
        if exclude is None:
            return False
        er0, er1, ec0, ec1 = exclude
        return er0 <= r <= er1 and ec0 <= c <= ec1

    others = [
        (r, c)
        for r, c in (other_legend_anchors or [])
        if not (r == anchor_row and c == anchor_col)
    ]

    def hits_other(r: int, c: int) -> bool:
        return any(r == or_ and c == oc for or_, oc in others)

    start_row, end_row = anchor_row, min(max_row, anchor_row + LEGEND_MIN_EXTRA_ROWS)
    start_col, end_col = anchor_col, min(max_col, anchor_col + LEGEND_MIN_EXTRA_COLS)

    for r, c in others:
        if r == anchor_row and c > anchor_col:
            end_col = min(end_col, c - 1)
        if c == anchor_col and r > anchor_row:
            end_row = min(end_row, r - 1)

    for _ in range(50):
        grew = False
        if end_row < max_row and (end_row - start_row + 1) < LEGEND_MAX_ROWS:
            r = end_row + 1
            row_texts = [_cell_str(ws, r, c) for c in range(start_col, end_col + 1)]
            if any(_is_legend_junk_neighbor(t) for t in row_texts):
                pass
            elif not any(hits_other(r, c) for c in range(start_col, end_col + 1)) and any(
                _cell_interesting(ws, r, c) and not in_exclude(r, c)
                for c in range(start_col, end_col + 1)
            ):
                end_row = r
                grew = True
        if end_col < max_col and (end_col - start_col + 1) < LEGEND_MAX_COLS:
            c = end_col + 1
            col_texts = [_cell_str(ws, r, c) for r in range(start_row, end_row + 1)]
            if any(_is_legend_junk_neighbor(t) for t in col_texts):
                pass
            elif not any(hits_other(r, c) for r in range(start_row, end_row + 1)) and any(
                _cell_interesting(ws, r, c) and not in_exclude(r, c)
                for r in range(start_row, end_row + 1)
            ):
                end_col = c
                grew = True
        if start_col > min_col and (end_col - start_col + 1) < LEGEND_MAX_COLS:
            c = start_col - 1
            col_texts = [_cell_str(ws, r, c) for r in range(start_row, end_row + 1)]
            if any(_is_legend_junk_neighbor(t) for t in col_texts):
                pass
            elif not any(hits_other(r, c) for r in range(start_row, end_row + 1)) and any(
                _cell_interesting(ws, r, c) and not in_exclude(r, c)
                for r in range(start_row, end_row + 1)
            ):
                start_col = c
                grew = True
        if not grew:
            break

    if end_row < start_row or end_col < start_col:
        return None
    if (end_row - start_row + 1) * (end_col - start_col + 1) > 400:
        return None
    return RangeBlock(
        label=f"legend@{get_column_letter(start_col)}{start_row}",
        start_row=start_row,
        end_row=end_row,
        start_col=start_col,
        end_col=end_col,
    )


def _picture_boxes(slide) -> list[tuple[int, int, int, int]]:
    boxes = [
        (int(s.left), int(s.top), int(s.width), int(s.height))
        for s in slide.shapes
        if s.shape_type == MSO_SHAPE_TYPE.PICTURE
    ]
    boxes.sort(key=lambda b: b[2] * b[3], reverse=True)
    return boxes


def _resolve_placement_boxes(
    slide,
    element: dict,
) -> tuple[
    tuple[int, int, int, int],
    tuple[int, int, int, int],
    list[tuple[int, int, int, int]],
]:
    """Return summary_box, metrics_box, legend_boxes.

    Legends always use compact left-margin slots (template legend pictures are often
    oversized half-slide boxes that stretch/overlap when filled).
    """
    if element.get("summary_box") and element.get("metrics_box"):
        summary = tuple(int(x) for x in element["summary_box"])
        metrics = tuple(int(x) for x in element["metrics_box"])
    else:
        summary = NORTH_SUMMARY_BOX
        metrics = NORTH_METRICS_BOX

    if element.get("legend_boxes"):
        legend_boxes = [tuple(int(x) for x in box) for box in element["legend_boxes"]]
    else:
        legend_boxes = list(NORTH_LEGEND_BOXES)

    slots = _picture_boxes(slide)
    if len(slots) >= 4:
        large = sorted(slots[:2], key=lambda b: (b[1], b[0]))
        summary, metrics = large[0], large[1]
        # Keep compact legend boxes unless caller overrode them.
        if not element.get("legend_boxes"):
            legend_boxes = list(NORTH_LEGEND_BOXES)
    elif len(slots) >= 1:
        # Largest slot (or only slot) is the main content area — split for tables.
        large = max(slots, key=lambda b: b[2] * b[3])
        left, top, width, height = large
        sum_h = int(height * 0.36)
        met_h = height - sum_h - _GAP
        summary = (left, top, width, sum_h)
        metrics = (left, top + sum_h + _GAP, width, max(met_h, 1))

    return summary, metrics, legend_boxes[:2]


def _cluster_hits_by_column(
    hits: list[tuple[int, int]],
    *,
    min_prefer: int = 3,
) -> list[tuple[int, int]]:
    """Keep hits from the densest column so legend-adjacent junk (OT/Hours) is ignored."""
    if not hits:
        return []
    by_col: dict[int, list[tuple[int, int]]] = {}
    for hit in hits:
        by_col.setdefault(hit[1], []).append(hit)
    best_col, cluster = max(by_col.items(), key=lambda kv: (len(kv[1]), -kv[0]))
    if len(cluster) >= min_prefer or len(by_col) == 1:
        return sorted(cluster)
    # Fall back to all hits if nothing is dense enough.
    return sorted(hits)


def detect_north_summary_layout(workbook_bytes: bytes, sheet_name: str) -> NorthSummaryLayout:
    """Detect summary table, metrics table, and two legend tables."""
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    match = _find_sheet_name(list(wb.sheetnames), sheet_name)
    if match is None:
        wb.close()
        raise ValueError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[match]
    min_row, max_row, min_col, max_col = _used_bounds(ws)

    labels = _find_row_labels(ws, min_row, max_row, min_col, max_col)
    category_hits = _cluster_hits_by_column(
        [labels[t] for t in CATEGORY_ROWS if t in labels],
        min_prefer=3,
    )
    kpi_hits = _cluster_hits_by_column(
        [labels[k] for k in KPI_KEYS if k in labels],
        min_prefer=4,
    )

    legend_anchors = [
        (r, c)
        for r in range(min_row, max_row + 1)
        for c in range(min_col, max_col + 1)
        if _is_legend_header(_cell_str(ws, r, c))
    ]

    # --- Summary table (GSE MPR + category rows) ---
    if category_hits:
        cat_rows = [r for r, _ in category_hits]
        cat_cols = [c for _, c in category_hits]
        near = min(cat_rows)
        header = _find_entity_header_row(
            ws, min_row, max_row, min_col, max_col, near_row=near
        )
        header_row = header[0] if header else None
        entity_end = header[2] if header else None
        entity_start = header[1] if header else None
        seed_cols = list(cat_cols)
        if entity_start is not None:
            seed_cols.append(entity_start)
        if "gse mpr" in labels:
            seed_cols.append(labels["gse mpr"][1])
            cat_rows.append(labels["gse mpr"][0])
        # Stop before first KPI row / second entity header so summary stays separate.
        # Only use clustered KPI rows (ignore OT/Hours junk matched as KPI aliases).
        stop: set[int] = set()
        if kpi_hits:
            first_kpi = min(r for r, _ in kpi_hits)
            stop.add(first_kpi)
            for row in range(max(r for r, _ in category_hits) + 1, first_kpi):
                ents = [
                    c
                    for c in range(min_col, max_col + 1)
                    if _is_entity_header(_cell_str(ws, row, c))
                ]
                if len(ents) >= 2:
                    stop.add(row)
                    break
        summary = _expand_table(
            ws,
            cat_rows,
            seed_cols,
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            header_row=header_row,
            entity_end_col=entity_end,
            stop_before_rows=stop,
        )
        # Hard cap: never past the row before the first clustered KPI label.
        if kpi_hits:
            summary.end_row = min(summary.end_row, min(r for r, _ in kpi_hits) - 1)
            if stop:
                summary.end_row = min(summary.end_row, min(stop) - 1)
        summary.label = "summary"
    else:
        summary = RangeBlock("summary", min_row, min_row, min_col, max_col)

    # --- Metrics table (KPI rows) ---
    if kpi_hits:
        kpi_rows = [r for r, _ in kpi_hits]
        kpi_cols = [c for _, c in kpi_hits]
        near = min(kpi_rows)
        header = _find_entity_header_row(
            ws, min_row, max_row, min_col, max_col, near_row=near
        )
        # Prefer a header at/above the KPI block (may repeat under summary).
        header_row = header[0] if header else None
        entity_end = header[2] if header else summary.end_col
        entity_start = header[1] if header else summary.start_col
        # If header sits inside the summary block, look for one nearer KPIs.
        if header_row is not None and header_row <= summary.end_row:
            alt = _find_entity_header_row(
                ws,
                summary.end_row + 1,
                max_row,
                min_col,
                max_col,
                near_row=near,
            )
            if alt:
                header_row, entity_start, entity_end = alt
            else:
                # Reuse summary columns; start at first KPI (no second header).
                header_row = None
                entity_end = summary.end_col
                entity_start = summary.start_col
        seed_cols = list(kpi_cols)
        if entity_start is not None:
            seed_cols.append(entity_start)
        metrics = _expand_table(
            ws,
            kpi_rows,
            seed_cols,
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            header_row=header_row,
            entity_end_col=entity_end,
            stop_before_rows={r for r, _ in legend_anchors if r > max(kpi_rows)},
        )
        metrics.label = "metrics"
        # Align metrics under the summary table so left legend / OT-Hours cols stay out.
        metrics.start_col = summary.start_col
        metrics.end_col = max(metrics.end_col, summary.end_col)
        for key in ("qc compliance", "budget $000s", "overtime", "total hours", "lead input"):
            if key in labels and any(labels[key][1] == c for _, c in kpi_hits):
                metrics.end_row = max(metrics.end_row, labels[key][0])
    else:
        metrics = RangeBlock(
            "metrics",
            summary.end_row + 1,
            max_row,
            summary.start_col,
            summary.end_col,
        )

    # Combined exclude so legends stay outside both tables.
    exclude = (
        min(summary.start_row, metrics.start_row),
        max(summary.end_row, metrics.end_row),
        min(summary.start_col, metrics.start_col),
        max(summary.end_col, metrics.end_col),
    )

    # Prefer exact known legend body labels so both full tables are captured:
    #   Legend + Score Above 4 / Between 3 and 4 / Below 3
    #   Legend + KPI Better/Worse than Goal / Not Applicable / Goal Pending
    score_hits = _find_legend_row_cells(
        ws,
        SCORE_LEGEND_ROWS,
        min_row,
        max_row,
        min_col,
        max_col,
        aliases=SCORE_LEGEND_ALIASES,
        exclude=exclude,
    )
    kpi_hits_legend = _find_legend_row_cells(
        ws,
        KPI_LEGEND_ROWS,
        min_row,
        max_row,
        min_col,
        max_col,
        aliases=KPI_LEGEND_ALIASES,
        exclude=exclude,
    )
    ordered: list[RangeBlock] = []
    score_block = _legend_block_from_rows(
        ws,
        kind="score",
        body_hits=score_hits,
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        exclude=exclude,
    )
    kpi_block = _legend_block_from_rows(
        ws,
        kind="kpi",
        body_hits=kpi_hits_legend,
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        exclude=exclude,
    )
    if score_block:
        ordered.append(score_block)
    if kpi_block:
        ordered.append(kpi_block)

    # Fallback: expand from "Legend" headers if exact labels were not found.
    if len(ordered) < 2:
        seen = {
            (b.start_row, b.end_row, b.start_col, b.end_col) for b in ordered
        }
        for row, col in sorted(legend_anchors, key=lambda rc: (rc[0], rc[1])):
            required_end = None
            # If this header sits above known body rows, force those rows in.
            body_below = [
                r
                for r, _c, _t in (score_hits + kpi_hits_legend)
                if r > row
            ]
            if body_below:
                required_end = max(body_below)
            block = _expand_legend_table(
                ws,
                row,
                col,
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                exclude=(
                    exclude
                    if (row > exclude[1] or col < exclude[2] or col > exclude[3])
                    else None
                ),
                other_legend_anchors=legend_anchors,
            )
            if block is None or (block.row_count < 2 and block.col_count < 2):
                continue
            if required_end is not None:
                block.end_row = max(block.end_row, required_end)
            key = (block.start_row, block.end_row, block.start_col, block.end_col)
            if key in seen:
                continue
            if any(
                abs(block.start_row - s[0]) <= 1 and abs(block.start_col - s[2]) <= 1
                for s in seen
            ):
                continue
            seen.add(key)
            kind = _legend_kind(
                ws, block.start_row, block.end_row, block.start_col, block.end_col
            )
            block.label = f"legend_{kind}"
            ordered.append(block)
            if len(ordered) >= 2:
                break

    # Prefer score legend first, KPI legend second (top/bottom on the slide).
    score_legends = [b for b in ordered if "score" in b.label]
    kpi_legends = [b for b in ordered if "kpi" in b.label]
    other_legends = [b for b in ordered if b not in score_legends and b not in kpi_legends]
    final: list[RangeBlock] = []
    final.extend(score_legends[:1])
    final.extend(kpi_legends[:1])
    for b in other_legends + score_legends[1:] + kpi_legends[1:]:
        if len(final) >= 2:
            break
        if b not in final:
            final.append(b)
    if len(final) < 2:
        final = sorted(ordered, key=lambda b: (b.start_row, b.start_col))[:2]
    ordered = final

    missing_score = [t for t in SCORE_LEGEND_ROWS if not any(t == h[2] for h in score_hits)]
    missing_kpi = [t for t in KPI_LEGEND_ROWS if not any(t == h[2] for h in kpi_hits_legend)]
    if missing_score or missing_kpi:
        logger.warning(
            "North legends missing labels score=%s kpi=%s",
            missing_score,
            missing_kpi,
        )
        print(
            f">>> WARNING North legends missing labels "
            f"score={missing_score or 'ok'} kpi={missing_kpi or 'ok'}"
        )
    else:
        print(
            ">>> North legends found all expected rows: "
            "Score Above/Between/Below 4 + KPI Better/Worse/N/A/Goal Pending"
        )

    wb.close()
    missing = [k for k in KPI_KEYS if k not in labels]
    if missing:
        logger.warning("North metrics missing KPI labels: %s", missing)

    print(
        f">>> North layout {match}: summary={summary.range_address} "
        f"metrics={metrics.range_address} "
        f"legends={[b.range_address for b in ordered]} "
        f"score_rows={[h[2] for h in score_hits]} "
        f"kpi_legend_rows={[h[2] for h in kpi_hits_legend]}"
    )
    return NorthSummaryLayout(summary=summary, metrics=metrics, legends=ordered)


def _capture_block(
    workbook_bytes: bytes,
    sheet_name: str,
    block: RangeBlock,
    *,
    prefer_com: bool,
) -> bytes | None:
    try:
        png = capture_range_png(
            workbook_bytes,
            sheet_name,
            block.start_row,
            block.end_row,
            block.start_col,
            block.end_col,
            prefer_excel_com=prefer_com,
        )
        return _validate_capture(
            png,
            label=f"north {block.label} {sheet_name}!{block.range_address}",
            min_w=30,
            min_h=15,
            require_wide=False,
        )
    except Exception as exc:
        msg = str(exc).casefold()
        if "mostly blank" in msg or "too small" in msg:
            try:
                from scorecard_screenshots import _capture_via_render, _png_from_pil

                raw = _capture_via_render(
                    workbook_bytes,
                    sheet_name,
                    block.start_row,
                    block.end_row,
                    block.start_col,
                    block.end_col,
                )
                with Image.open(io.BytesIO(raw)) as img:
                    if img.width >= 30 and img.height >= 15:
                        return _png_from_pil(img)
            except Exception:
                pass
        logger.warning("North capture failed for %s: %s", block.range_address, exc)
        return None


def apply_north_summary_panels(slide, data, element: dict) -> bool:
    """Screenshot Visualizations Scorecard Summaries onto North Scorecard Summary."""
    workbook = element.get("workbook", "visualizations")
    prefer_com = bool(element.get("prefer_excel_com", True))
    fit = str(element.get("fit", "fill")).lower()
    replace_existing = bool(element.get("replace_existing_pictures", True))

    try:
        workbook_bytes = data.store.workbook_bytes(workbook)
    except FileNotFoundError as exc:
        logger.warning("North summary skipped: %s", exc)
        print(f">>> North summary skipped — visualizations workbook missing ({exc})")
        return False

    available = []
    try:
        available = list(data.sheet_names(workbook))
    except Exception:
        available = []

    try:
        sheet_name = resolve_sheet_name(
            workbook_bytes,
            sheet=element.get("sheet"),
            sheet_index=element.get("sheet_index"),
            sheet_match=element.get("sheet_match", ["scorecard summar"]),
            sheet_match_index=int(element.get("sheet_match_index", 0) or 0),
            available=available or None,
        )
    except Exception as exc:
        logger.warning("Could not resolve Scorecard Summaries sheet: %s", exc)
        return False

    try:
        layout = detect_north_summary_layout(workbook_bytes, sheet_name)
    except Exception as exc:
        logger.warning("North summary layout detection failed: %s", exc)
        return False

    summary_png = _capture_block(
        workbook_bytes, sheet_name, layout.summary, prefer_com=prefer_com
    )
    metrics_png = _capture_block(
        workbook_bytes, sheet_name, layout.metrics, prefer_com=prefer_com
    )
    legend_pngs: list[bytes] = []
    for legend in layout.legends:
        png = _capture_block(workbook_bytes, sheet_name, legend, prefer_com=prefer_com)
        if png:
            legend_pngs.append(png)

    if not summary_png and not metrics_png:
        print(f">>> FAILED north summary capture from {workbook}/{sheet_name}")
        return False

    summary_box, metrics_box, legend_boxes = _resolve_placement_boxes(slide, element)

    if replace_existing:
        removed = _remove_slide_pictures(slide)
        if removed:
            print(f">>> Removed {removed} template picture(s) for North summary")

    out_dir = Path(getattr(data.store, "base_dir", Path("."))) / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    placed = 0

    def _place(
        png: bytes,
        box: tuple[int, int, int, int],
        name: str,
        addr: str,
        *,
        fit_mode: str | None = None,
        align: str = "center",
    ) -> None:
        nonlocal placed
        place_picture_on_slide(
            slide,
            png,
            left=box[0],
            top=box[1],
            max_width=box[2],
            max_height=box[3],
            fit=fit_mode or fit,
            align=align,
        )
        placed += 1
        try:
            with Image.open(io.BytesIO(png)) as img:
                debug = out_dir / f"_debug_north_{name}_{img.width}x{img.height}.png"
                img.save(debug)
                print(
                    f">>> North {name} from {sheet_name}!{addr} "
                    f"({img.width}x{img.height}) box={box} fit={fit_mode or fit} -> {debug.name}"
                )
        except Exception:
            print(f">>> North {name} placed from {sheet_name}!{addr}")

    if summary_png:
        _place(summary_png, summary_box, "summary", layout.summary.range_address)
    if metrics_png:
        _place(metrics_png, metrics_box, "metrics", layout.metrics.range_address)

    # Legends: preserve Excel proportions (contain) so they stay compact like the template.
    legend_fit = str(element.get("legend_fit", "contain")).lower()
    for idx, png in enumerate(legend_pngs[:2]):
        box = legend_boxes[idx] if idx < len(legend_boxes) else NORTH_LEGEND_BOXES[min(idx, 1)]
        legend = layout.legends[idx]
        _place(
            png,
            box,
            f"legend{idx + 1}",
            legend.range_address,
            fit_mode=legend_fit,
            align="top-left",
        )

    if len(legend_pngs) < 2:
        print(
            f">>> WARNING: expected 2 legend screenshots, captured {len(legend_pngs)} "
            f"from {sheet_name}"
        )

    print(
        f"\n>>> Slide 14 North Scorecard Summary: placed {placed} screenshot(s) "
        f"(summary + metrics + legends) from {workbook}/{sheet_name}\n"
    )
    return placed > 0
