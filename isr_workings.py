"""ISR Motorized slide (PPT 12) from New GSE MPR Workings.xlsx → ISR tab.

Only these pieces from the ISR tab:
  1. Regions | RELIABILITY | SEVERITY table (MTD/YTD/Score through SYSTEM)
  2. Reliability and Severity Excel graphs

Leading Issues / Action Plans are cleared to empty editable text boxes.
"""

from __future__ import annotations

import io
import logging
import re
import tempfile
import time
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from pptx.enum.shapes import MSO_SHAPE_TYPE

from gir_panels import clear_leading_action_narrative
from people_workings import (
    _com_chart_title,
    _screenshot_excel_chart,
)
from scorecard_screenshots import (
    _find_com_worksheet,
    _find_sheet_name,
    _open_excel_workbook,
    _range_address,
    _validate_capture,
    capture_range_png,
    place_picture_on_slide,
    resolve_sheet_name,
)

logger = logging.getLogger(__name__)

# Match the ISR MOTORIZED template layout:
#   - Regions | RELIABILITY | SEVERITY table in the upper content slot
#   - Reliability / Severity graphs sized like the template charts below
ISR_TABLE_BOX = (313_417, 1_056_453, 8_014_153, 2_840_631)

ISR_CHART_BOXES = {
    "reliability": (372_129, 4_299_857, 4_167_214, 1_794_522),
    "severity": (4_647_937, 4_299_857, 3_576_773, 1_792_224),
}

ISR_CHART_SPECS = [
    {"key": "reliability", "title": "Reliability", "match": ["reliability", "rel"]},
    {"key": "severity", "title": "Severity", "match": ["severity", "sev"]},
]

_ROW_END_LABELS = {"system"}
_STOP_LABELS = (
    "weight",
    "kpi",
    "total score",
    "leading issue",
    "action plan",
    "notes",
    "month",
    "jan",
    "overall total",
    "motorized",
    "stationary",
)


def _cell_str(ws, row: int, col: int) -> str:
    val = ws.cell(row, col).value
    if val is None:
        return ""
    return str(val).strip()


def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").casefold()).strip()


def _row_blob(ws, row: int, max_col: int) -> str:
    return " ".join(_norm(_cell_str(ws, row, c)) for c in range(1, max_col + 1))


def _merged_max_col(ws, row: int, col: int) -> int:
    try:
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                return int(merged.max_col)
    except Exception:
        pass
    return col


def _is_month_header(text: str) -> bool:
    t = _norm(text)
    if not t:
        return False
    months = {
        "jan", "feb", "mar", "apr", "may", "jun",
        "jul", "aug", "sep", "oct", "nov", "dec", "ytd",
    }
    if t in months:
        return True
    return t[:3] in months and len(t) <= 9


def _find_isr_regions_header(ws, *, max_row: int = 40, max_col: int = 40) -> tuple[int, int, int] | None:
    """Find Regions + RELIABILITY + SEVERITY header (two sections only).

    Returns (header_row, start_col, end_col). Never includes monthly
    Actual/Goal chart-source grids to the right of SEVERITY.
    """
    scan_cols = min(max(int(ws.max_column or max_col), 20), max_col)
    for row in range(1, min(int(ws.max_row or max_row), max_row) + 1):
        blob = _row_blob(ws, row, scan_cols)
        if "region" not in blob:
            continue
        has_rel = "reliability" in blob
        has_sev = "severity" in blob
        next_blob = _row_blob(ws, row + 1, scan_cols)
        has_mtd = "mtd" in blob or "mtd" in next_blob or "actual" in next_blob
        if not has_mtd or not has_rel or not has_sev:
            continue

        start_col = None
        rel_col = None
        sev_col = None
        for col in range(1, scan_cols + 1):
            text = _norm(_cell_str(ws, row, col))
            if not text:
                continue
            if start_col is None and "region" in text:
                start_col = col
            if "reliability" in text or text == "rel":
                rel_col = col if rel_col is None else min(rel_col, col)
            if "severity" in text or text == "sev":
                sev_col = col if sev_col is None else min(sev_col, col)

        if start_col is None or sev_col is None:
            continue

        # Severity metric block is ~6 cols (Actual .. YTD Score).
        end_col = max(_merged_max_col(ws, row, sev_col), sev_col + 5)

        for r in (row, row + 1, row + 2):
            for col in range(start_col, min(scan_cols, end_col + 2) + 1):
                text = _cell_str(ws, r, col)
                if not text:
                    continue
                if _is_month_header(text):
                    end_col = min(end_col, col - 1)
                    break
                low = _norm(text)
                if low in {"reliability", "severity"} and col > sev_col + 5:
                    end_col = min(end_col, col - 1)
                    break
                if col <= sev_col + 5:
                    end_col = max(end_col, col)

        end_col = min(end_col, sev_col + 6)
        end_col = max(end_col, sev_col + 5)

        print(
            f">>> ISR Regions header row {row}: "
            f"RELIABILITY={'col' + str(rel_col) if rel_col else 'n/a'}, "
            f"SEVERITY=col{sev_col} → cols {start_col}-{end_col} "
            f"(two sections only; monthly grids excluded)"
        )
        return row, start_col, end_col
    return None


def _discover_isr_table(workbook_bytes: bytes, sheet_name: str) -> tuple[int, int, int, int]:
    """Locate ONLY the Regions RELIABILITY + SEVERITY table (two sections)."""
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    try:
        match = _find_sheet_name(list(wb.sheetnames), sheet_name) or sheet_name
        ws = wb[match]
        hdr = _find_isr_regions_header(ws)
        if hdr is None:
            raise ValueError(
                f"Could not find Regions RELIABILITY / SEVERITY table on {sheet_name!r}"
            )

        start_row, start_col, end_col = hdr
        header_rows = 1
        for extra in (1, 2):
            blob = _row_blob(ws, start_row + extra, end_col)
            if any(token in blob for token in ("mtd", "ytd", "score", "actual", "goal")):
                header_rows = extra + 1
            else:
                break

        end_row = start_row + header_rows - 1
        max_scan = min(int(ws.max_row or 80), start_row + 40)
        blank_streak = 0
        saw_system = False
        for row in range(start_row + header_rows, max_scan + 1):
            label = _norm(_cell_str(ws, row, start_col))
            row_has = any(_cell_str(ws, row, col) for col in range(start_col, end_col + 1))
            if not row_has:
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue
            blank_streak = 0

            blob = _row_blob(ws, row, end_col)
            if any(token in label for token in _STOP_LABELS) or any(
                token in blob for token in ("weight", "total score")
            ):
                break

            end_row = row
            if label in _ROW_END_LABELS:
                saw_system = True
                break

        if end_row < start_row + header_rows:
            end_row = min(max_scan, start_row + header_rows + 12)

        # Do not extend into monthly chart-source grids.
        addr = _range_address(start_row, end_row, start_col, end_col)
        logger.info("ISR Regions table %s!%s", match, addr)
        print(
            f">>> ISR table range: {match}!{addr} "
            f"(Regions / RELIABILITY / SEVERITY only"
            f"{', through SYSTEM' if saw_system else ''}; "
            f"monthly Actual/Goal grids not included)"
        )
        return start_row, end_row, start_col, end_col
    finally:
        wb.close()


def _match_isr_chart(title: str) -> dict | None:
    lower = (title or "").casefold()
    # Prefer exact-ish matches; avoid accidental "rel" inside unrelated words later via order.
    for spec in ISR_CHART_SPECS:
        for token in spec["match"]:
            if token == "rel" and "reliability" not in lower and not re.search(r"\brel\b", lower):
                continue
            if token == "sev" and "severity" not in lower and not re.search(r"\bsev\b", lower):
                continue
            if token in lower:
                return spec
    return None


def _iter_sheet_charts(ws) -> list[tuple[float, float, str, object]]:
    found: list[tuple[float, float, str, object]] = []
    seen: set[str] = set()

    def _add(obj, *, source: str) -> None:
        try:
            name = str(getattr(obj, "Name", "") or "")
        except Exception:
            name = ""
        key = name.casefold() or f"{source}:{id(obj)}"
        if key in seen:
            return
        seen.add(key)
        try:
            top = float(getattr(obj, "Top", len(found) * 100))
            left = float(getattr(obj, "Left", len(found) * 100))
        except Exception:
            top, left = float(len(found) * 100), 0.0
        title = _com_chart_title(obj) or name
        found.append((top, left, title, obj))
        print(f">>> Found ISR Excel graph ({source}): {title or 'untitled'!r}")

    try:
        count = int(ws.ChartObjects().Count)
    except Exception:
        count = 0
    for idx in range(1, count + 1):
        try:
            _add(ws.ChartObjects(idx), source=f"ChartObjects[{idx}]")
        except Exception as exc:
            logger.warning("ISR ChartObjects(%s) unreadable: %s", idx, exc)

    try:
        shape_count = int(ws.Shapes.Count)
    except Exception:
        shape_count = 0
    for idx in range(1, shape_count + 1):
        try:
            shape = ws.Shapes(idx)
            if int(getattr(shape, "Type", 0) or 0) != 3:
                continue
            _add(shape, source=f"Shapes[{idx}]")
        except Exception:
            continue

    found.sort(key=lambda item: (item[0], item[1]))
    return found


def _screenshot_isr_charts(workbook_path: Path, sheet_name: str) -> dict[str, bytes]:
    """Screenshot only Reliability + Severity graphs from the ISR sheet."""
    excel = None
    wb = None
    by_key: dict[str, bytes] = {}
    ordered: list[bytes] = []
    try:
        excel, wb = _open_excel_workbook(workbook_path)
        ws = _find_com_worksheet(wb, sheet_name)
        ws.Activate()
        try:
            excel.ActiveWindow.Zoom = 100
        except Exception:
            pass
        time.sleep(0.3)

        charts = _iter_sheet_charts(ws)
        print(f">>> ISR sheet: {len(charts)} Excel graph(s) found")
        scored: list[tuple[float, float, str, bytes]] = []
        for idx, (top, left, title, chart_obj) in enumerate(charts, start=1):
            lower = (title or "").casefold()
            if any(
                token in lower
                for token in ("motorized", "stationary", "people", "finance", "non-motor", "pmi")
            ):
                print(f">>> Skipping non Rel/Sev graph {idx}: {title!r}")
                continue
            label = f"ISR graph {idx} ({title or 'untitled'})"
            try:
                try:
                    chart_obj.Activate()
                except Exception:
                    try:
                        chart_obj.Select()
                    except Exception:
                        pass
                time.sleep(0.2)
                data = _screenshot_excel_chart(chart_obj, label=label)
                scored.append((top, left, title, data))
                print(f">>> Screenshotted ISR graph {idx}: {title!r} ({len(data)} bytes)")
            except Exception as exc:
                logger.warning("%s failed: %s", label, exc)
                print(f">>> WARNING: could not screenshot ISR graph {idx}: {exc}")

        scored.sort(key=lambda item: (item[0], item[1]))
        for top, left, title, data in scored:
            spec = _match_isr_chart(title)
            if spec and spec["key"] not in by_key:
                by_key[spec["key"]] = data
                print(f">>> Matched ISR screenshot to '{spec['title']}' (title {title!r})")
            else:
                ordered.append(data)

        for spec in ISR_CHART_SPECS:
            if spec["key"] in by_key:
                continue
            if not ordered:
                break
            by_key[spec["key"]] = ordered.pop(0)
            print(f">>> Assigned position-ordered ISR screenshot to '{spec['title']}'")
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
    return by_key


def _remove_isr_content_shapes(slide) -> int:
    """Remove prior table pictures and native charts (keep footer logo)."""
    removed = 0
    for shape in list(slide.shapes):
        drop = False
        st = shape.shape_type
        if st in (
            MSO_SHAPE_TYPE.EMBEDDED_OLE_OBJECT,
            getattr(MSO_SHAPE_TYPE, "LINKED_OLE_OBJECT", MSO_SHAPE_TYPE.EMBEDDED_OLE_OBJECT),
        ):
            drop = True
        elif getattr(shape, "has_chart", False):
            drop = True
        elif st == MSO_SHAPE_TYPE.PICTURE and 900_000 <= int(shape.top) < 6_200_000:
            drop = True
        if not drop:
            continue
        shape._element.getparent().remove(shape._element)
        removed += 1
    return removed


def apply_isr_workings_panels(slide, data, element: dict) -> bool:
    """Fill ISR Motorized from Workings!ISR Regions Rel/Sev table + graphs."""
    workbook = element.get("workbook", "workings")
    prefer_com = bool(element.get("prefer_excel_com", True))
    # Table fills the template OLE-style slot; graphs keep template chart sizes.
    fit = str(element.get("fit", "fill")).lower()
    chart_fit = str(element.get("chart_fit", "fill")).lower()

    try:
        workbook_bytes = data.store.workbook_bytes(workbook)
    except FileNotFoundError as exc:
        logger.warning("ISR workings screenshot skipped: %s", exc)
        return False

    available = []
    try:
        available = list(data.sheet_names(workbook))
    except Exception:
        available = []

    try:
        sheet_name = resolve_sheet_name(
            workbook_bytes,
            sheet=element.get("sheet", "ISR"),
            sheet_index=element.get("sheet_index"),
            sheet_match=element.get("sheet_match", ["ISR", "ISR MOTORIZED"]),
            sheet_match_index=int(element.get("sheet_match_index", 0) or 0),
            available=available or None,
        )
    except Exception as exc:
        logger.warning("Could not resolve Workings ISR sheet: %s", exc)
        return False

    out_dir = Path(getattr(data.store, "base_dir", Path("."))) / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    placed = 0

    table_png = None
    try:
        start_row, end_row, start_col, end_col = _discover_isr_table(workbook_bytes, sheet_name)
        table_png = capture_range_png(
            workbook_bytes,
            sheet_name,
            start_row,
            end_row,
            start_col,
            end_col,
            prefer_excel_com=prefer_com,
        )
        table_png = _validate_capture(
            table_png,
            label=f"ISR Regions table {sheet_name}",
            min_w=200,
            min_h=80,
            require_wide=True,
        )
    except Exception as exc:
        logger.warning("ISR Regions table capture failed: %s", exc)
        print(f">>> ERROR: ISR Regions RELIABILITY/SEVERITY table screenshot failed: {exc}")
        table_png = None

    charts_by_key: dict[str, bytes] = {}
    if prefer_com:
        try:
            with tempfile.TemporaryDirectory(prefix="mpr_isr_") as tmp:
                path = Path(tmp) / "workings.xlsx"
                path.write_bytes(workbook_bytes)
                charts_by_key = _screenshot_isr_charts(path, sheet_name)
        except Exception as exc:
            logger.error("ISR Excel graph screenshots unavailable: %s", exc)
            print(f">>> ERROR: could not screenshot ISR graphs from Excel: {exc}")
    else:
        print(">>> ERROR: ISR graphs require Excel COM screenshots (prefer_excel_com=true)")

    if not table_png and not charts_by_key:
        logger.warning("No ISR screenshots from workings/%s", sheet_name)
        return False

    removed = _remove_isr_content_shapes(slide)
    logger.info("Slide 12 cleared %s picture/chart/OLE shape(s)", removed)

    if table_png:
        left, top, width, height = ISR_TABLE_BOX
        place_picture_on_slide(
            slide,
            table_png,
            left=left,
            top=top,
            max_width=width,
            max_height=height,
            fit=fit,
        )
        placed += 1
        try:
            with Image.open(io.BytesIO(table_png)) as img:
                debug = out_dir / f"_debug_isr_table_{img.width}x{img.height}.png"
                img.save(debug)
                print(
                    f">>> ISR Regions (RELIABILITY + SEVERITY) table screenshot placed "
                    f"from workings/{sheet_name} ({img.width}x{img.height}) -> {debug.name}"
                )
        except Exception:
            print(f">>> ISR Regions table screenshot placed from workings/{sheet_name}")

    for spec in ISR_CHART_SPECS:
        png = charts_by_key.get(spec["key"])
        if png is None:
            print(f">>> WARNING: missing Excel screenshot for ISR '{spec['title']}' graph")
            continue
        left, top, width, height = ISR_CHART_BOXES[spec["key"]]
        place_picture_on_slide(
            slide,
            png,
            left=left,
            top=top,
            max_width=width,
            max_height=height,
            fit=chart_fit,
        )
        placed += 1
        try:
            with Image.open(io.BytesIO(png)) as img:
                debug = out_dir / f"_debug_isr_{spec['key']}_{img.width}x{img.height}.png"
                img.save(debug)
                print(
                    f">>> ISR '{spec['title']}' graph screenshot placed "
                    f"({img.width}x{img.height}, fit={chart_fit}) -> {debug.name}"
                )
        except Exception:
            print(f">>> ISR '{spec['title']}' graph screenshot placed")

    if bool(element.get("clear_narrative", True)):
        n = clear_leading_action_narrative(slide)
        print(f">>> ISR Leading Issues / Action Plans cleared ({n} text box(es))")

    print(
        f"\n>>> Slide 12 ISR Motorized: placed {placed} screenshot(s) from workings/{sheet_name} "
        f"(Rel/Sev table={'yes' if table_png else 'no'}, graphs={len(charts_by_key)}/2)\n"
    )
    return placed > 0
