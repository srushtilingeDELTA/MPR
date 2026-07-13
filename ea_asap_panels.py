"""EA Compliance / ASAP slide from New GSE MPR Workings.xlsx.

The Workings dashboard is already one combined EA COMPLIANCE | ASAP REPORTING
table. Capture that table once and place it once into the template OLE slot on
slide 6 (do not stitch multiple sheet screenshots — that duplicates the table).
"""

from __future__ import annotations

import io
import logging
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from pptx.enum.shapes import MSO_SHAPE_TYPE

from gir_panels import clear_leading_action_narrative
from scorecard_screenshots import (
    _find_sheet_name,
    _norm_sheet_key,
    _validate_capture,
    capture_sheet_png,
    place_picture_on_slide,
    resolve_sheet_name,
)

logger = logging.getLogger(__name__)

# Left content slot on slide 6 — larger than the template OLE so the table reads clearly,
# while staying left of Leading Issues (~8.42M) and above the footer logo (~6.35M).
EA_CONTENT_LEFT = 260_000
EA_CONTENT_TOP = 820_000
EA_CONTENT_WIDTH = 8_050_000
EA_CONTENT_HEIGHT = 4_650_000

# Prefer a single sheet that already hosts the combined table.
PREFERRED_SHEET_MATCHES = [
    ["EAC", "ASAP"],
    ["EA COMPLIANCE", "ASAP"],
    ["EA", "ASAP"],
    ["EAC"],
    ["EA COMPLIANCE"],
    ["EA COMPLIANCE / ASAP"],
    ["ASAP"],
]


def _remove_left_content_media(slide) -> int:
    """Remove OLE/pictures in the left content area; keep logo + right narrative."""
    removed = 0
    content_right = EA_CONTENT_LEFT + EA_CONTENT_WIDTH + 200_000
    content_bottom = EA_CONTENT_TOP + EA_CONTENT_HEIGHT + 200_000
    for shape in list(slide.shapes):
        st = shape.shape_type
        is_ole = st in (
            MSO_SHAPE_TYPE.EMBEDDED_OLE_OBJECT,
            getattr(MSO_SHAPE_TYPE, "LINKED_OLE_OBJECT", MSO_SHAPE_TYPE.EMBEDDED_OLE_OBJECT),
        )
        is_pic = st == MSO_SHAPE_TYPE.PICTURE
        if not is_ole and not is_pic:
            continue
        left = int(shape.left)
        top = int(shape.top)
        # Keep footer logo under the content slot.
        if top >= content_bottom:
            continue
        # Keep anything that lives in the Leading Issues column.
        if left >= content_right:
            continue
        if top + int(shape.height) < EA_CONTENT_TOP - 100_000:
            continue
        shape._element.getparent().remove(shape._element)
        removed += 1
        logger.info("Removed slide 6 content shape %s (%s)", shape.name, st)
    return removed


def _sheet_mentions(workbook_bytes: bytes, sheet_name: str, tokens: list[str]) -> int:
    """Count how many tokens appear in the first block of the sheet (for ranking)."""
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
    try:
        match = _find_sheet_name(list(wb.sheetnames), sheet_name)
        if match is None:
            return 0
        ws = wb[match]
        blob_parts: list[str] = []
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=40, max_col=30, values_only=True)):
            for val in row:
                if val is None:
                    continue
                blob_parts.append(str(val).casefold())
            if idx > 40:
                break
        blob = " | ".join(blob_parts)
        score = 0
        for token in tokens:
            if token.casefold() in blob:
                score += 1
        # Strong bonus when both EA and ASAP headers are present on one sheet.
        if "ea compliance" in blob and "asap" in blob:
            score += 5
        if "eac" in blob and "asap" in blob:
            score += 3
        return score
    finally:
        wb.close()


def _resolve_single_ea_asap_sheet(
    workbook_bytes: bytes,
    available: list[str],
    element: dict,
) -> str:
    """Pick exactly one Workings sheet for the combined EA/ASAP table."""
    # Explicit override wins.
    if element.get("sheet") or element.get("sheet_match") or element.get("sheet_index") is not None:
        return resolve_sheet_name(
            workbook_bytes,
            sheet=element.get("sheet"),
            sheet_index=element.get("sheet_index"),
            sheet_match=element.get("sheet_match"),
            sheet_match_index=int(element.get("sheet_match_index", 0) or 0),
            available=available or None,
        )

    ranked: list[tuple[int, str]] = []
    for patterns in PREFERRED_SHEET_MATCHES:
        try:
            name = resolve_sheet_name(
                workbook_bytes,
                sheet_match=patterns,
                available=available or None,
            )
        except Exception:
            continue
        score = _sheet_mentions(workbook_bytes, name, patterns)
        # Prefer exact-ish names.
        key = _norm_sheet_key(name)
        if key in {"eac", "asap", "ea compliance", "ea compliance / asap", "eac / asap"}:
            score += 2
        if "eac" in key and "asap" in key:
            score += 4
        ranked.append((score, name))

    if not ranked:
        raise ValueError(
            "No EAC/ASAP sheet found in Workings. "
            f"Available: {available}"
        )

    ranked.sort(
        key=lambda item: (
            -item[0],
            0 if "eac" in _norm_sheet_key(item[1]) and "asap" not in _norm_sheet_key(item[1]) else 1,
            item[1],
        )
    )
    best_score, best_name = ranked[0]
    logger.info("Selected Workings sheet %r for EA/ASAP (score=%s)", best_name, best_score)
    return best_name


def apply_ea_asap_workings_panels(slide, data, element: dict) -> bool:
    """Screenshot one Workings EA/ASAP table and place it once in the template slot."""
    workbook = element.get("workbook", "workings")
    prefer_com = bool(element.get("prefer_excel_com", True))
    # Fill the OLE rectangle like the template embedded object.
    fit = str(element.get("fit", "fill")).lower()

    try:
        workbook_bytes = data.store.workbook_bytes(workbook)
    except FileNotFoundError as exc:
        logger.warning("EA/ASAP workings screenshot skipped: %s", exc)
        return False

    available = []
    try:
        available = list(data.sheet_names(workbook))
    except Exception:
        available = []

    try:
        sheet_name = _resolve_single_ea_asap_sheet(workbook_bytes, available, element)
    except Exception as exc:
        logger.warning("Could not resolve Workings EA/ASAP sheet: %s", exc)
        return False

    try:
        png = capture_sheet_png(
            workbook_bytes,
            sheet_name,
            prefer_excel_com=prefer_com,
            max_rows=int(element.get("max_rows", 80) or 80),
            max_cols=int(element.get("max_cols", 30) or 30),
        )
        png = _validate_capture(
            png,
            label=f"Workings {sheet_name}",
            min_w=80,
            min_h=40,
            require_wide=False,
        )
    except Exception as exc:
        logger.warning("Capture failed for Workings!%s: %s", sheet_name, exc)
        return False

    out_dir = Path(getattr(data.store, "base_dir", Path("."))) / "output"
    try:
        with Image.open(io.BytesIO(png)) as img:
            out_dir.mkdir(parents=True, exist_ok=True)
            debug = out_dir / f"_debug_ea_asap_{img.width}x{img.height}.png"
            img.save(debug)
            print(
                f">>> EA/ASAP captured once from workings/{sheet_name} "
                f"({img.width}x{img.height}) -> {debug.name}"
            )
    except Exception:
        print(f">>> EA/ASAP captured once from workings/{sheet_name}")

    removed = _remove_left_content_media(slide)
    logger.info("Slide 6 cleared %s OLE/picture shape(s) before placing EA/ASAP", removed)

    place_picture_on_slide(
        slide,
        png,
        left=EA_CONTENT_LEFT,
        top=EA_CONTENT_TOP,
        max_width=EA_CONTENT_WIDTH,
        max_height=EA_CONTENT_HEIGHT,
        fit=fit,
    )
    print(
        f"\n>>> Slide 6 EA/ASAP: placed 1 table screenshot from workings/{sheet_name} "
        f"into template OLE slot ({EA_CONTENT_WIDTH}x{EA_CONTENT_HEIGHT} EMU, fit={fit})\n"
    )

    if bool(element.get("clear_narrative", True)):
        n = clear_leading_action_narrative(slide)
        print(f">>> EA/ASAP Leading Issues / Action Plan cleared ({n} text box(es))")

    return True
