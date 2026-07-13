"""Inspect the GIR tab in New GSE MPR Workings.xlsx.

Usage:
    python scripts\\inspect_workings_gir.py
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from report_utils import load_config
from workbook_store import WorkbookStore


def main() -> int:
    config = load_config(base_dir=BASE_DIR)
    store = WorkbookStore(config, BASE_DIR)
    try:
        store.load()
        raw = store.workbook_bytes("workings")
    except FileNotFoundError as exc:
        print(exc)
        print("Sync SharePoint first, or place New GSE MPR Workings.xlsx under data/.")
        return 1

    wb = load_workbook(io.BytesIO(raw), data_only=True)
    print("Workings sheets:")
    for idx, name in enumerate(wb.sheetnames):
        mark = " <== GIR" if "gir" in name.strip().casefold() else ""
        print(f"  [{idx}] {name!r}{mark}")

    match = next((n for n in wb.sheetnames if n.strip().casefold() == "gir"), None)
    if match is None:
        match = next((n for n in wb.sheetnames if "gir" in n.strip().casefold()), None)
    if match is None:
        print("\nNo GIR sheet found.")
        wb.close()
        return 1

    ws = wb[match]
    print(f"\nSheet {match!r} used range preview (first 40 non-empty rows, cols A-P):")
    shown = 0
    for row in range(1, min(120, (ws.max_row or 1) + 1)):
        parts = []
        for col in range(1, 17):
            val = ws.cell(row, col).value
            if val is None or str(val).strip() == "":
                continue
            parts.append(f"{get_column_letter(col)}={str(val).strip()[:40]!r}")
        if parts:
            print(f"  R{row}: " + " | ".join(parts))
            shown += 1
            if shown >= 40:
                break
    wb.close()
    print("\nSlide 5 fills native PowerPoint tables/chart from Actuals.")
    print("Injury Breakdown may be filled from Workings!GIR when that table parses cleanly.")
    print("Leading Issues / Action Plan text boxes are cleared for manual entry.")
    print("Screenshots are not used for GIR.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
