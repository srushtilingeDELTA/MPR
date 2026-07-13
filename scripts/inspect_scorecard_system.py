"""Inspect / debug System tab section detection for scorecard screenshots.

Usage:
    python scripts/inspect_scorecard_system.py
    python scripts/inspect_scorecard_system.py --dump-left
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

import io

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from report_utils import load_config
from scorecard_screenshots import detect_system_layout, select_sections_for_slide
from workbook_store import WorkbookStore


def _fmt(r1: int, r2: int, c1: int, c2: int) -> str:
    return f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"


def _dump_left(raw: bytes, sheet_name: str = "System", max_rows: int = 120) -> None:
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    match = next((n for n in wb.sheetnames if n.strip().casefold() == sheet_name.casefold()), None)
    if match is None:
        print(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
        return
    ws = wb[match]
    print(f"\nLeft columns dump ({match}), rows 1-{max_rows}, cols A-F:")
    print(f"Merged ranges (first 30): {list(ws.merged_cells.ranges)[:30]}")
    for row in range(1, max_rows + 1):
        parts = []
        for col in range(1, 7):
            val = ws.cell(row, col).value
            if val is None or str(val).strip() == "":
                continue
            parts.append(f"{get_column_letter(col)}={str(val).strip()[:40]!r}")
        if parts:
            print(f"  R{row}: " + " | ".join(parts))
    wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Inspect System scorecard sections.")
    parser.add_argument(
        "--dump-left",
        action="store_true",
        help="Print left-column values to debug category detection",
    )
    args = parser.parse_args()

    config = load_config(base_dir=BASE_DIR)
    store = WorkbookStore(config, BASE_DIR)
    try:
        store.load()
        raw = store.workbook_bytes("scorecards")
    except FileNotFoundError as exc:
        print(exc)
        print("Sync SharePoint files or place the workbook under data/ first.")
        return 1

    if args.dump_left:
        _dump_left(raw)

    layout = detect_system_layout(raw, sheet_name="System")
    sections = layout.sections
    if not sections:
        print("No sections detected on System sheet.")
        return 1

    print(
        f"Header rows {layout.header_start_row}-{layout.header_end_row}  "
        f"cols {get_column_letter(layout.start_col)}-{get_column_letter(layout.end_col)}"
    )
    print(f"Found {len(sections)} category section(s) on System:\n")
    for section in sections:
        flag = " [BLACK]" if section.is_black else ""
        print(
            f"  {section.index + 1}. {section.title!r}  "
            f"rows {section.start_row}-{section.end_row}{flag}"
        )

    slide3 = select_sections_for_slide(sections, mode="first", count=3)
    slide4 = select_sections_for_slide(sections, mode="last", count=2, include_black=True)
    b3 = layout.capture_bounds_for(slide3)
    b4 = layout.capture_bounds_for(slide4)
    print("\nSlide 3 (PPT) capture:")
    print(
        f"  {_fmt(layout.header_start_row, b3[1], b3[2], b3[3])}  -> "
        + ", ".join(s.title for s in slide3)
    )
    print("\nSlide 4 (PPT) capture:")
    if layout.needs_header_stitch(slide4):
        print(
            f"  header {_fmt(layout.header_start_row, layout.header_end_row, b4[2], b4[3])} "
            f"+ body {_fmt(b4[0], b4[1], b4[2], b4[3])}  -> "
            + ", ".join(s.title for s in slide4)
        )
    else:
        print(
            f"  {_fmt(layout.header_start_row, b4[1], b4[2], b4[3])}  -> "
            + ", ".join(s.title for s in slide4)
        )

    if len(sections) < 2:
        print(
            "\nWARNING: only one section detected. Re-run with --dump-left and share the output "
            "so category splits can be tuned to your workbook."
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
