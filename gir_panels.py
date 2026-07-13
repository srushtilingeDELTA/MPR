"""GIR slide panels from New GSE MPR Workings.xlsx → GIR tab.

Captures multiple Excel regions (not the raw ALL / NO 115 dumps) and places
them into the template GIR slots so the slide matches the designed layout.
"""

from __future__ import annotations

import io
import logging
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image

from ppt_format import clear_text_frame_content, set_text_frame_preserve
from scorecard_screenshots import (
    _find_sheet_name,
    _open_excel_workbook,
    _find_com_worksheet,
    _png_from_pil,
    _remove_slide_pictures,
    _remove_slide_tables_and_charts,
    _validate_capture,
    capture_range_png,
    place_picture_on_slide,
    resolve_sheet_name,
    _range_address,
)

logger = logging.getLogger(__name__)

# Template GIR data-band slots (EMU) from GSE MPR - Template.pptx slide 5.
GIR_CHART_BOX = (423264, 995240, 5084907, 1650332)
GIR_SUMMARY_BOX = (5747723, 812128, 1657704, 1833444)  # May / YoY
GIR_INJURY_BOX = (7840215, 812130, 3876698, 1833443)  # Injury Breakdown
GIR_RECORDABLE_BOX = (423263, 2732550, 11293650, 741697)
GIR_METRIC_BOX = (423263, 3645516, 11293650, 790575)

# Clear everything from under the title down to Leading Issues.
GIR_CLEAR_TOP = 700_000
GIR_CLEAR_BOTTOM = 4_550_000


@dataclass
class ExcelBlock:
    name: str
    start_row: int
    end_row: int
    start_col: int
    end_col: int

    @property
    def address(self) -> str:
        return _range_address(self.start_row, self.end_row, self.start_col, self.end_col)


def _cell_str(ws: Worksheet, row: int, col: int) -> str:
    val = ws.cell(row, col).value
    if val is None:
        return ""
    return str(val).strip()


def _find_token_cell(
    ws: Worksheet,
    tokens: list[str],
    *,
    max_row: int = 80,
    max_col: int = 40,
    exact: bool = False,
) -> tuple[int, int] | None:
    needles = [t.casefold() for t in tokens]
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            text = _cell_str(ws, row, col).casefold()
            if not text:
                continue
            for needle in needles:
                if exact and text == needle:
                    return row, col
                if not exact and needle in text:
                    return row, col
    return None


def _dashboard_end_col(ws: Worksheet) -> int:
    """Column just left of the raw ALL / NO 115 / Yr_Nb dumps."""
    markers = ("yr_nb", "all", "no 115", "no115", "no. 115")
    for row in range(1, 40):
        for col in range(1, min(50, int(ws.max_column or 50)) + 1):
            text = _cell_str(ws, row, col).casefold()
            if text in markers or text.startswith("no 115"):
                return max(1, col - 1)
    # Fallback: first ~14 columns are usually the formatted dashboard.
    return min(14, int(ws.max_column or 14))


def _expand_block(
    ws: Worksheet,
    header_row: int,
    header_col: int,
    *,
    max_row: int,
    max_col: int,
    min_rows: int = 2,
    min_cols: int = 2,
) -> ExcelBlock:
    """Grow a table block from a header cell until empty border."""
    end_col = header_col
    for col in range(header_col + 1, max_col + 1):
        if not _cell_str(ws, header_row, col) and not any(
            _cell_str(ws, r, col) for r in range(header_row, min(header_row + 8, max_row + 1))
        ):
            break
        end_col = col

    end_row = header_row
    empty_streak = 0
    for row in range(header_row + 1, max_row + 1):
        if any(_cell_str(ws, row, c) for c in range(header_col, end_col + 1)):
            end_row = row
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak >= 2 and (end_row - header_row + 1) >= min_rows:
                break

    end_row = max(end_row, header_row + min_rows - 1)
    end_col = max(end_col, header_col + min_cols - 1)
    return ExcelBlock("block", header_row, end_row, header_col, end_col)


def discover_gir_blocks(workbook_bytes: bytes, sheet_name: str) -> dict[str, ExcelBlock]:
    """Locate dashboard blocks on the GIR tab (excluding raw data dumps)."""
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    match = _find_sheet_name(list(wb.sheetnames), sheet_name)
    if match is None:
        wb.close()
        raise ValueError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[match]
    max_row = min(int(ws.max_row or 80), 100)
    dash_end = _dashboard_end_col(ws)
    max_col = dash_end
    logger.info("GIR dashboard columns A-%s (raw dumps excluded)", get_column_letter(dash_end))

    blocks: dict[str, ExcelBlock] = {}

    injury = _find_token_cell(
        ws, ["injury breakdown", "nonrec", "non-rec", "dart"], max_row=max_row, max_col=max_col
    )
    if injury:
        # Prefer a header row that says Injury Breakdown; else use DART/NonRec row.
        hdr = _find_token_cell(ws, ["injury breakdown"], max_row=max_row, max_col=max_col)
        row, col = hdr or injury
        # Injury table often starts a column or two left of the match.
        start_col = max(1, col - 1)
        block = _expand_block(ws, row, start_col, max_row=max_row, max_col=max_col, min_rows=4, min_cols=4)
        # Include a title row above when present.
        if row > 1 and "injur" in _cell_str(ws, row - 1, start_col).casefold():
            block = ExcelBlock("injury", row - 1, block.end_row, block.start_col, block.end_col)
        else:
            block = ExcelBlock("injury", block.start_row, block.end_row, block.start_col, block.end_col)
        blocks["injury"] = block

    recordable = _find_token_cell(ws, ["recordable"], max_row=max_row, max_col=max_col)
    if recordable:
        row, col = recordable
        block = _expand_block(ws, row, max(1, col), max_row=max_row, max_col=max_col, min_rows=2, min_cols=10)
        blocks["recordable"] = ExcelBlock(
            "recordable", block.start_row, block.end_row, 1, max_col
        )

    metric = _find_token_cell(ws, ["month to date", "metric"], max_row=max_row, max_col=max_col)
    if metric:
        row, col = metric
        # Metric header is often one row above "Month to Date".
        start_row = row
        if "month to date" in _cell_str(ws, row, col).casefold() and row > 1:
            start_row = row - 1
        block = _expand_block(ws, start_row, 1, max_row=max_row, max_col=max_col, min_rows=2, min_cols=8)
        blocks["metric"] = ExcelBlock("metric", block.start_row, block.end_row, 1, max_col)

    # May / YoY summary boxes — look for Actual/Plan near a month label.
    summary = _find_token_cell(ws, ["yo1y", "yo2y", "yoy actuals", "yoy"], max_row=max_row, max_col=max_col)
    if summary:
        row, col = summary
        start_row = max(1, row - 4)
        start_col = max(1, col - 1)
        end_row = min(max_row, row + 3)
        end_col = min(max_col, col + 2)
        blocks["summary"] = ExcelBlock("summary", start_row, end_row, start_col, end_col)
    else:
        actual = _find_token_cell(ws, ["actual:"], max_row=max_row, max_col=max_col)
        if actual:
            row, col = actual
            blocks["summary"] = ExcelBlock(
                "summary", max(1, row - 1), min(max_row, row + 5), max(1, col - 1), min(max_col, col + 2)
            )

    # Top-left chart/grid area: from sheet top to just above Recordable, left of summary/injury.
    rec_start = blocks["recordable"].start_row if "recordable" in blocks else max_row
    top_end = max(1, rec_start - 1)
    chart_end_col = max_col
    if "summary" in blocks:
        chart_end_col = min(chart_end_col, max(1, blocks["summary"].start_col - 1))
    if "injury" in blocks:
        chart_end_col = min(chart_end_col, max(1, blocks["injury"].start_col - 1))
    if chart_end_col >= 3 and top_end >= 3:
        blocks["chart"] = ExcelBlock("chart", 1, top_end, 1, chart_end_col)

    # If we failed to split the top, keep one combined top dashboard band.
    if "chart" not in blocks and "summary" not in blocks and "injury" not in blocks and top_end >= 2:
        blocks["top"] = ExcelBlock("top", 1, top_end, 1, max_col)

    wb.close()
    for name, block in blocks.items():
        logger.info("GIR block %-12s %s", name, block.address)
    return blocks


def _export_excel_charts(workbook_path: Path, sheet_name: str) -> list[bytes]:
    """Export embedded Excel charts on the GIR sheet as PNGs."""
    excel = None
    wb = None
    pngs: list[bytes] = []
    try:
        excel, wb = _open_excel_workbook(workbook_path)
        ws = _find_com_worksheet(wb, sheet_name)
        ws.Activate()
        count = int(ws.ChartObjects().Count)
        for idx in range(1, count + 1):
            chart_obj = ws.ChartObjects(idx)
            export_path = Path(tempfile.gettempdir()) / f"mpr_gir_chart_{idx}_{Path(workbook_path).stem}.png"
            try:
                if export_path.exists():
                    export_path.unlink()
                chart_obj.Chart.Export(str(export_path))
                if export_path.exists() and export_path.stat().st_size > 1500:
                    data = export_path.read_bytes()
                    try:
                        data = _validate_capture(data, label=f"GIR chart {idx}", min_w=120, min_h=80)
                    except Exception:
                        # Charts can be narrower; still usable.
                        with Image.open(io.BytesIO(data)) as img:
                            data = _png_from_pil(img)
                    pngs.append(data)
                    logger.info("Exported GIR Excel chart %s (%s bytes)", idx, len(data))
            except Exception as exc:
                logger.info("GIR chart %s export failed: %s", idx, exc)
            finally:
                try:
                    if export_path.exists():
                        export_path.unlink()
                except Exception:
                    pass
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
    return pngs


def _capture_block_png(
    workbook_bytes: bytes,
    sheet_name: str,
    block: ExcelBlock,
    *,
    prefer_excel_com: bool = True,
) -> bytes:
    return capture_range_png(
        workbook_bytes,
        sheet_name,
        block.start_row,
        block.end_row,
        block.start_col,
        block.end_col,
        prefer_excel_com=prefer_excel_com,
    )


def clear_leading_action_narrative(slide) -> int:
    """Clear Leading Issues / Action Plan body text; keep headers and empty editable boxes.

    Works across GIR (bottom band) and EA/ASAP (right-side stacked boxes).
    """
    headers = []
    for shape in slide.shapes:
        if not getattr(shape, "has_text_frame", False):
            continue
        text = (shape.text_frame.text or "").strip().casefold()
        if "leading issue" in text or text.startswith("action plan"):
            headers.append(shape)

    # Slide 6 sometimes has an empty Action Plan title box above the body.
    if headers:
        for shape in slide.shapes:
            if not getattr(shape, "has_text_frame", False):
                continue
            # Only fill empty title text boxes — not the red header rectangles.
            if "textbox" not in (shape.name or "").casefold():
                continue
            text = (shape.text_frame.text or "").strip()
            if text:
                continue
            top = int(shape.top)
            if int(shape.height) > 350_000 or int(shape.height) < 200_000:
                continue
            for header in headers:
                if abs(int(shape.left) - int(header.left)) < 800_000 and top > int(header.top) + 1_500_000:
                    set_text_frame_preserve(shape.text_frame, "Action Plan")
                    headers.append(shape)
                    logger.info("Restored empty Action Plan header on %s", shape.name)
                    break

    cleared = 0
    for shape in slide.shapes:
        if not getattr(shape, "has_text_frame", False):
            continue
        text = (shape.text_frame.text or "").strip()
        lower = text.casefold()
        if "leading issue" in lower or lower.startswith("action plan"):
            continue
        if int(shape.height) < 500_000:
            continue
        if int(shape.top) < 800_000:
            continue

        mid_x = int(shape.left) + int(shape.width) // 2
        near_header = False
        for header in headers:
            h_mid = int(header.left) + int(header.width) // 2
            # Same column as a Leading Issues / Action Plan header, and below it.
            if abs(mid_x - h_mid) <= 2_200_000 and int(shape.top) > int(header.top) + 100_000:
                near_header = True
                break
        if not near_header and not headers:
            # Fallback for GIR bottom band when headers weren't matched.
            if int(shape.top) >= 4_850_000:
                near_header = True
        if not near_header:
            continue

        clear_text_frame_content(shape.text_frame)
        tf = shape.text_frame
        if tf.paragraphs:
            para = tf.paragraphs[0]
            if para.runs:
                para.runs[0].text = ""
            else:
                para.add_run().text = ""
        cleared += 1
        logger.info("Cleared narrative box %s", shape.name)

    return cleared


def clear_gir_narrative_textboxes(slide) -> int:
    """Backward-compatible alias for GIR slide narrative clearing."""
    return clear_leading_action_narrative(slide)


def apply_gir_workings_panels(slide, data, element: dict) -> bool:
    """Build the GIR slide from multiple Workings!GIR screenshots into template slots."""
    workbook = element.get("workbook", "workings")
    prefer_com = bool(element.get("prefer_excel_com", True))
    fit = str(element.get("fit", "fill")).lower()

    try:
        workbook_bytes = data.store.workbook_bytes(workbook)
    except FileNotFoundError as exc:
        logger.warning("GIR workings screenshot skipped: %s", exc)
        return False

    available = []
    try:
        available = list(data.sheet_names(workbook))
    except Exception:
        available = []

    try:
        sheet_name = resolve_sheet_name(
            workbook_bytes,
            sheet=element.get("sheet", "GIR"),
            sheet_index=element.get("sheet_index"),
            sheet_match=element.get("sheet_match", ["GIR"]),
            sheet_match_index=int(element.get("sheet_match_index", 0) or 0),
            available=available or None,
        )
    except Exception as exc:
        logger.warning("GIR sheet resolve failed: %s", exc)
        return False

    # Clear old template data shapes in the content band.
    _remove_slide_pictures(slide)
    removed = _remove_slide_tables_and_charts(
        slide,
        remove_tables=True,
        remove_charts=True,
        remove_band_labels=True,
        band_top=GIR_CLEAR_TOP,
        band_bottom=GIR_CLEAR_BOTTOM,
    )
    logger.info("GIR slide cleared %s table/chart/label shape(s)", removed)

    if bool(element.get("clear_narrative", True)):
        n = clear_gir_narrative_textboxes(slide)
        print(f">>> GIR Leading Issues / Action Plan cleared ({n} text box(es))")

    # Discover Excel blocks (override via yaml ranges when provided).
    overrides = element.get("ranges") or {}
    blocks = discover_gir_blocks(workbook_bytes, sheet_name)
    for key, addr in overrides.items():
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(str(addr))
        blocks[key] = ExcelBlock(key, min_row, max_row, min_col, max_col)

    placed = 0
    out_dir = Path(getattr(data.store, "base_dir", Path("."))) / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    def _place(name: str, png: bytes, box: tuple[int, int, int, int]) -> None:
        nonlocal placed
        left, top, width, height = box
        try:
            png = _validate_capture(png, label=f"GIR {name}", min_w=80, min_h=40, require_wide=False)
        except Exception as exc:
            logger.warning("GIR panel %s rejected: %s", name, exc)
            return
        place_picture_on_slide(
            slide,
            png,
            left=left,
            top=top,
            max_width=width,
            max_height=height,
            fit=fit,
        )
        try:
            with Image.open(io.BytesIO(png)) as img:
                debug = out_dir / f"_debug_gir_{name}_{img.width}x{img.height}.png"
                img.save(debug)
                print(f">>> GIR panel '{name}' placed ({img.width}x{img.height}) -> {debug.name}")
        except Exception:
            print(f">>> GIR panel '{name}' placed ({len(png)} bytes)")
        placed += 1

    # 1) Prefer a real Excel chart for the System GIR chart slot.
    chart_pngs: list[bytes] = []
    if prefer_com:
        try:
            with tempfile.TemporaryDirectory(prefix="mpr_gir_") as tmp:
                path = Path(tmp) / "workings.xlsx"
                path.write_bytes(workbook_bytes)
                chart_pngs = _export_excel_charts(path, sheet_name)
        except Exception as exc:
            logger.info("GIR Excel chart export unavailable: %s", exc)
    if chart_pngs:
        _place("chart", chart_pngs[0], GIR_CHART_BOX)
    elif "chart" in blocks:
        _place("chart", _capture_block_png(workbook_bytes, sheet_name, blocks["chart"], prefer_excel_com=prefer_com), GIR_CHART_BOX)
    elif "top" in blocks:
        _place("top", _capture_block_png(workbook_bytes, sheet_name, blocks["top"], prefer_excel_com=prefer_com), GIR_CHART_BOX)

    # 2) Summary + Injury Breakdown
    if "summary" in blocks:
        _place(
            "summary",
            _capture_block_png(workbook_bytes, sheet_name, blocks["summary"], prefer_excel_com=prefer_com),
            GIR_SUMMARY_BOX,
        )
    if "injury" in blocks:
        _place(
            "injury",
            _capture_block_png(workbook_bytes, sheet_name, blocks["injury"], prefer_excel_com=prefer_com),
            GIR_INJURY_BOX,
        )

    # 3) Recordable + Metric tables
    if "recordable" in blocks:
        _place(
            "recordable",
            _capture_block_png(workbook_bytes, sheet_name, blocks["recordable"], prefer_excel_com=prefer_com),
            GIR_RECORDABLE_BOX,
        )
    if "metric" in blocks:
        _place(
            "metric",
            _capture_block_png(workbook_bytes, sheet_name, blocks["metric"], prefer_excel_com=prefer_com),
            GIR_METRIC_BOX,
        )

    # Fallback: if discovery found almost nothing, capture left dashboard only (no raw dumps).
    if placed == 0:
        dash_end = 14
        try:
            wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
            match = _find_sheet_name(list(wb.sheetnames), sheet_name) or sheet_name
            dash_end = _dashboard_end_col(wb[match])
            wb.close()
        except Exception:
            pass
        fallback = ExcelBlock("dashboard", 1, 45, 1, dash_end)
        box = (314628, 720000, 11534806, 3800000)
        _place(
            "dashboard",
            _capture_block_png(workbook_bytes, sheet_name, fallback, prefer_excel_com=prefer_com),
            box,
        )

    print(f"\n>>> GIR slide: placed {placed} panel screenshot(s) from workings/{sheet_name}\n")
    return placed > 0
