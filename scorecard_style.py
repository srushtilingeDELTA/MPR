"""Apply GSE MPR template table styling (colors, headers, sections) from Excel blocks."""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

import pandas as pd

from ppt_format import merge_table_row, style_table_cell
from scorecard_data import (
    OPPORTUNITIES_PATTERN,
    OVERALL_TOTAL_SCORE_PATTERN,
    SCORECARD_ROW_TYPES,
    _detect_section,
    format_scorecard_cell,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Extracted from template embedded Excel (SCORECARD SUMMARIES / EA-ASAP tables).
HEADER_FILL = "003366"
HEADER_FONT = "FFFFFF"
DATA_FILL = "FFFFFF"
ALT_ROW_FILL = "F2F2F2"
SUBTOTAL_FILL = "D9D9D9"
GROUP_ROW_FILL = "7D9BC1"
GROUP_ROW_FONT = "FFFFFF"

SECTION_FILLS = {
    "Safety & Security": "991933",
    "Safety": "991933",
    "Customer Experience": "003366",
    "Operations": "EAAA00",
    "Finance": "8F993E",
    "People": "7D9BC1",
}

SECTION_FONT = {
    "Operations": "000000",
}

ROW_KIND_SECTION = "section"
ROW_KIND_COLUMN_HEADER = "column_header"
ROW_KIND_KPI = "kpi"
ROW_KIND_SUBROW = "subrow"
ROW_KIND_TOTAL_SCORE = "total_score"
ROW_KIND_OVERALL_TOTAL = "overall_total"
ROW_KIND_OPPORTUNITIES = "opportunities"
ROW_KIND_DATA = "data"

COMPLIANCE_GROUP_LABELS = {"MOTORIZED", "STATIONARY", "SYSTEM"}
COMPLIANCE_HEADER_ROWS = 3


def section_fill(section: str | None) -> str | None:
    if not section:
        return None
    for key, color in SECTION_FILLS.items():
        if key.lower() in section.lower() or section.lower() in key.lower():
            return color
    return None


def section_font_color(section: str | None) -> str:
    if section and section in SECTION_FONT:
        return SECTION_FONT[section]
    if section and "operations" in section.lower():
        return "000000"
    return HEADER_FONT


def _row_values(block: pd.DataFrame, row_idx: int) -> list[str]:
    return [format_scorecard_cell(block.iat[row_idx, col]) for col in range(len(block.columns))]


def classify_scorecard_row(
    values: list[str],
    *,
    current_section: str = "",
) -> tuple[str, str]:
    """Return (section_name, row_kind) for one scorecard table row."""
    joined = " ".join(v.strip() for v in values if v and v.strip())
    section = current_section

    detected = _detect_section(joined)
    if detected:
        return detected, ROW_KIND_SECTION

    weight = values[1].strip() if len(values) > 1 else ""
    label = values[2].strip() if len(values) > 2 else ""
    lowered = label.lower()

    if weight.upper() == "WEIGHT" and label.upper() == "KPI":
        return section, ROW_KIND_COLUMN_HEADER

    if lowered in SCORECARD_ROW_TYPES:
        return section, ROW_KIND_SUBROW

    if lowered == "total score":
        return section, ROW_KIND_TOTAL_SCORE

    if OVERALL_TOTAL_SCORE_PATTERN.search(label):
        return section, ROW_KIND_OVERALL_TOTAL

    if OPPORTUNITIES_PATTERN.search(label):
        return section, ROW_KIND_OPPORTUNITIES

    if label and lowered != "kpi":
        return section, ROW_KIND_KPI

    if weight and "%" in weight:
        return section, ROW_KIND_KPI

    return section, ROW_KIND_DATA


def _openpyxl_fill_hex(cell) -> str | None:
    fill = getattr(cell, "fill", None)
    if fill is None or fill.fill_type is None:
        return None
    fg = fill.fgColor
    if fg.type == "rgb" and fg.rgb:
        rgb = fg.rgb[-6:].upper()
        return rgb
    if fg.type == "theme" and fg.theme == 0:
        return DATA_FILL
    return None


def _openpyxl_font_bold(cell) -> bool | None:
    if cell.font and cell.font.bold is not None:
        return bool(cell.font.bold)
    return None


def _openpyxl_font_color(cell) -> str | None:
    if cell.font and cell.font.color and cell.font.color.type == "rgb" and cell.font.color.rgb:
        return cell.font.color.rgb[-6:].upper()
    return None


def style_scorecard_row(
    table,
    row_idx: int,
    *,
    section: str,
    kind: str,
    ncols: int,
    font_size_pt: float = 8,
) -> None:
    """Apply template colors/fonts to one scorecard row."""
    sec_fill = section_fill(section)
    sec_font = section_font_color(section)

    if kind == ROW_KIND_SECTION:
        fill = sec_fill or HEADER_FILL
        font = sec_font
        for col_idx in range(ncols):
            cell = table.cell(row_idx, col_idx)
            style_table_cell(
                cell,
                cell.text,
                fill_hex=fill,
                font_bold=True,
                font_color_hex=font,
                font_size_pt=min(font_size_pt + 1, MAX_FONT_PT),
            )
        merge_table_row(table, row_idx, 0, ncols - 1)
        return

    if kind == ROW_KIND_COLUMN_HEADER:
        for col_idx in range(ncols):
            cell = table.cell(row_idx, col_idx)
            style_table_cell(
                cell,
                cell.text,
                fill_hex=HEADER_FILL,
                font_bold=True,
                font_color_hex=HEADER_FONT,
                font_size_pt=font_size_pt,
            )
        return

    if kind == ROW_KIND_OVERALL_TOTAL:
        for col_idx in range(ncols):
            cell = table.cell(row_idx, col_idx)
            style_table_cell(
                cell,
                cell.text,
                fill_hex=HEADER_FILL,
                font_bold=True,
                font_color_hex=HEADER_FONT,
                font_size_pt=font_size_pt,
            )
        return

    if kind == ROW_KIND_TOTAL_SCORE:
        for col_idx in range(ncols):
            cell = table.cell(row_idx, col_idx)
            style_table_cell(
                cell,
                cell.text,
                fill_hex=SUBTOTAL_FILL,
                font_bold=True,
                font_size_pt=font_size_pt,
            )
        return

    if kind == ROW_KIND_KPI:
        for col_idx in range(ncols):
            cell = table.cell(row_idx, col_idx)
            fill = sec_fill if col_idx <= 2 and cell.text.strip() else DATA_FILL
            font = sec_font if col_idx <= 2 and fill == sec_fill else None
            style_table_cell(
                cell,
                cell.text,
                fill_hex=fill,
                font_bold=bool(font),
                font_color_hex=font,
                font_size_pt=font_size_pt,
            )
        return

    fill = ALT_ROW_FILL if kind == ROW_KIND_SUBROW and row_idx % 2 == 0 else DATA_FILL
    for col_idx in range(ncols):
        cell = table.cell(row_idx, col_idx)
        style_table_cell(cell, cell.text, fill_hex=fill, font_size_pt=font_size_pt)


MAX_FONT_PT = 9.0


def apply_scorecard_block_styles(table, block: pd.DataFrame, *, font_size_pt: float = 8) -> None:
    """Color-code a scorecard table to match the template screenshot layout."""
    if block.empty:
        return

    ncols = len(table.columns)
    section = ""
    for row_idx in range(min(len(table.rows), len(block))):
        values = _row_values(block, row_idx)
        if not any(v.strip() for v in values):
            continue
        section, kind = classify_scorecard_row(values, current_section=section)
        style_scorecard_row(
            table,
            row_idx,
            section=section,
            kind=kind,
            ncols=ncols,
            font_size_pt=font_size_pt,
        )


def apply_styles_from_worksheet(
    table,
    ws: Worksheet,
    *,
    row_offset: int = 0,
    col_offset: int = 0,
    nrows: int | None = None,
    ncols: int | None = None,
) -> None:
    """Mirror Excel cell fills/fonts onto an existing PowerPoint table."""
    nrows = nrows if nrows is not None else len(table.rows)
    ncols = ncols if ncols is not None else len(table.columns)
    for r in range(nrows):
        for c in range(ncols):
            ppt_cell = table.cell(r, c)
            xl_cell = ws.cell(row=row_offset + r + 1, column=col_offset + c + 1)
            fill = _openpyxl_fill_hex(xl_cell)
            bold = _openpyxl_font_bold(xl_cell)
            font_color = _openpyxl_font_color(xl_cell)
            if fill or bold is not None or font_color:
                style_table_cell(
                    ppt_cell,
                    ppt_cell.text,
                    fill_hex=fill,
                    font_bold=bold,
                    font_color_hex=font_color,
                    font_size_pt=8,
                )


def _compliance_row_label(table, row_idx: int) -> str:
    return table.cell(row_idx, 0).text.strip().upper()


def apply_compliance_table_styles(table) -> None:
    """Style EA/ASAP compliance table like the template screenshot."""
    ncols = len(table.columns)
    for row_idx in range(min(COMPLIANCE_HEADER_ROWS, len(table.rows))):
        for col_idx in range(ncols):
            cell = table.cell(row_idx, col_idx)
            style_table_cell(
                cell,
                cell.text,
                fill_hex=HEADER_FILL,
                font_bold=True,
                font_color_hex=HEADER_FONT,
                font_size_pt=8,
            )

    for row_idx in range(COMPLIANCE_HEADER_ROWS, len(table.rows)):
        label = _compliance_row_label(table, row_idx)
        if label in COMPLIANCE_GROUP_LABELS:
            fill = HEADER_FILL if label == "SYSTEM" else GROUP_ROW_FILL
            font = HEADER_FONT
            for col_idx in range(ncols):
                cell = table.cell(row_idx, col_idx)
                style_table_cell(
                    cell,
                    cell.text,
                    fill_hex=fill,
                    font_bold=True,
                    font_color_hex=font,
                    font_size_pt=8,
                )
        else:
            fill = ALT_ROW_FILL if row_idx % 2 == 0 else DATA_FILL
            for col_idx in range(ncols):
                cell = table.cell(row_idx, col_idx)
                style_table_cell(cell, cell.text, fill_hex=fill, font_size_pt=8)


def apply_worksheet_table_styles(
    table,
    ws: Worksheet,
    df: pd.DataFrame,
    *,
    origin_row: int,
    origin_col: int,
) -> None:
    """
    Style a pasted worksheet block: prefer Excel fills, then infer header/section rows.
    """
    apply_styles_from_worksheet(
        table,
        ws,
        row_offset=origin_row,
        col_offset=origin_col,
        nrows=len(df.index),
        ncols=len(df.columns),
    )

    ncols = len(table.columns)
    section = ""
    for row_idx in range(len(df.index)):
        values = [format_scorecard_cell(df.iat[row_idx, col]) for col in range(len(df.columns))]
        if not any(v.strip() for v in values):
            continue
        detected = _detect_section(" ".join(values))
        if detected:
            section = detected
            style_scorecard_row(table, row_idx, section=section, kind=ROW_KIND_SECTION, ncols=ncols)
            continue
        section_name, kind = classify_scorecard_row(values, current_section=section)
        if kind in {ROW_KIND_COLUMN_HEADER, ROW_KIND_SECTION, ROW_KIND_OVERALL_TOTAL}:
            style_scorecard_row(table, row_idx, section=section_name, kind=kind, ncols=ncols)
