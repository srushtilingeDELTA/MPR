"""North Scorecard Comparison (PPT 15) from GSE MPR Visualizations.xlsx.

Screenshots the Entity / Period KPI comparison table from the
Scorecard Comparison tab (May'26, vs LM, vs LYSM by station).
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from scorecard_screenshots import (
    _fill_rgb,
    _find_sheet_name,
    _range_address,
    _remove_slide_pictures,
    _used_bounds,
    _validate_capture,
    capture_range_png,
    place_picture_on_slide,
    resolve_sheet_name,
)

logger = logging.getLogger(__name__)

# Full-bleed content band under the North Scorecard Comparison title (EMU).
# Template slide 15 has no picture slot — place into this wide content area.
COMPARISON_LEFT = 314_628
COMPARISON_TOP = 720_000
COMPARISON_WIDTH = 11_534_806
COMPARISON_HEIGHT = 5_400_000

# Required / strong header signals for the North comparison grid.
REQUIRED_HEADERS = ("entity", "period")
KPI_HEADERS = (
    "gir",
    "eac",
    "asap",
    "isr",
    "pmi",
    "pmi(n)",
    "pmi n",
    "qc",
    "budget",
    "ot",
    "total hours",
    "lic",
)
PERIOD_MARKERS = ("vs lm", "vs lysm", "vs ly", "vs sm")


@dataclass
class ComparisonTable:
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    label: str = "comparison"

    @property
    def range_address(self) -> str:
        return _range_address(self.start_row, self.end_row, self.start_col, self.end_col)


def _cell_str(ws, row: int, col: int) -> str:
    val = ws.cell(row, col).value
    if val is None:
        return ""
    return str(val).strip()


def _norm(text: str) -> str:
    cleaned = (
        str(text or "")
        .replace("\xa0", " ")
        .replace("\u2007", " ")
        .replace("\u202f", " ")
        .replace("\r", " ")
        .replace("\n", " ")
        .replace("\t", " ")
    )
    cleaned = re.sub(r"[()]", " ", cleaned)
    return re.sub(r"\s+", " ", cleaned).casefold().strip()


def _header_tokens(ws, row: int, min_col: int, max_col: int) -> dict[str, int]:
    """Map normalized header text → column for one row."""
    found: dict[str, int] = {}
    for col in range(min_col, max_col + 1):
        text = _norm(_cell_str(ws, row, col))
        if not text:
            continue
        found[text] = col
        # Also store compact form without spaces (pmi(n) / totalhours).
        compact = text.replace(" ", "")
        found.setdefault(compact, col)
    return found


def _score_header_row(tokens: dict[str, int]) -> int:
    """Higher is better — require Entity + Period, prefer KPI columns."""
    keys = set(tokens)
    if "entity" not in keys or "period" not in keys:
        return -1
    score = 20
    for kpi in KPI_HEADERS:
        if kpi in keys or kpi.replace(" ", "") in keys:
            score += 3
    return score


def _row_has_signal(ws, row: int, start_col: int, end_col: int) -> bool:
    for col in range(start_col, end_col + 1):
        if _cell_str(ws, row, col):
            return True
        try:
            rgb = _fill_rgb(ws.cell(row, col))
        except Exception:
            rgb = None
        if rgb is not None and not (rgb[0] >= 245 and rgb[1] >= 245 and rgb[2] >= 245):
            return True
    return False


def _looks_like_period_row(ws, row: int, period_col: int, start_col: int, end_col: int) -> bool:
    period = _norm(_cell_str(ws, row, period_col))
    if period:
        if any(m in period for m in PERIOD_MARKERS):
            return True
        # Month labels like May'26 / May 2026 / 5/26
        if re.search(r"[a-z]{3}.{0,3}\d{2,4}", period) or re.search(r"\d{1,2}[/.-]\d{2,4}", period):
            return True
    # Entity-merged rows may only have fills / KPI values.
    return _row_has_signal(ws, row, start_col, end_col)


def detect_scorecard_comparison_table(
    workbook_bytes: bytes,
    sheet_name: str,
) -> ComparisonTable:
    """Locate the Entity/Period comparison grid on Scorecard Comparison."""
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    try:
        match = _find_sheet_name(list(wb.sheetnames), sheet_name) or sheet_name
        ws = wb[match]
        min_row, max_row, min_col, max_col = _used_bounds(ws)

        best_row = None
        best_score = -1
        best_tokens: dict[str, int] = {}
        scan_end = min(max_row, min_row + 80)
        for row in range(min_row, scan_end + 1):
            tokens = _header_tokens(ws, row, min_col, max_col)
            score = _score_header_row(tokens)
            if score > best_score:
                best_score = score
                best_row = row
                best_tokens = tokens

        if best_row is None or best_score < 0:
            raise ValueError(
                f"Could not find Entity/Period comparison header on {match!r}. "
                f"Used range {_range_address(min_row, max_row, min_col, max_col)}"
            )

        entity_col = best_tokens["entity"]
        period_col = best_tokens["period"]
        # Columns: Entity through last populated header on that row.
        header_cols = sorted(set(best_tokens.values()))
        start_col = min(entity_col, period_col, header_cols[0])
        end_col = max(header_cols)
        # Extend a little right if adjacent KPI cells have header-ish fills/text.
        for col in range(end_col + 1, min(max_col, end_col + 6) + 1):
            if _cell_str(ws, best_row, col) or (
                _fill_rgb(ws.cell(best_row, col)) is not None
            ):
                end_col = col
            else:
                break

        # Body: continue while period/entity/KPI rows have content.
        end_row = best_row
        blank_streak = 0
        for row in range(best_row + 1, min(max_row, best_row + 60) + 1):
            if _looks_like_period_row(ws, row, period_col, start_col, end_col):
                end_row = row
                blank_streak = 0
                continue
            blank_streak += 1
            if blank_streak >= 2:
                break

        if end_row <= best_row:
            raise ValueError(
                f"Comparison header at row {best_row} on {match!r} has no data rows"
            )

        table = ComparisonTable(
            start_row=best_row,
            end_row=end_row,
            start_col=start_col,
            end_col=end_col,
            label="north_comparison",
        )
        print(
            f">>> North comparison table: {match}!{table.range_address} "
            f"({table.end_row - table.start_row + 1} rows x "
            f"{table.end_col - table.start_col + 1} cols)"
        )
        return table
    finally:
        wb.close()


def _capture_comparison_png(
    workbook_bytes: bytes,
    sheet_name: str,
    table: ComparisonTable,
    *,
    prefer_com: bool,
) -> bytes:
    png = capture_range_png(
        workbook_bytes,
        sheet_name,
        table.start_row,
        table.end_row,
        table.start_col,
        table.end_col,
        prefer_excel_com=prefer_com,
        zoom=125,
        min_width=1800,
    )
    return _validate_capture(
        png,
        label=f"north comparison {sheet_name}!{table.range_address}",
        min_w=80,
        min_h=40,
        require_wide=True,
    )


def apply_scorecard_comparison(slide, data, element: dict) -> bool:
    """Screenshot Visualizations Scorecard Comparison onto North Scorecard Comparison."""
    workbook = element.get("workbook", "visualizations")
    prefer_com = bool(element.get("prefer_excel_com", True))
    replace_existing = bool(element.get("replace_existing_pictures", True))
    fit = str(element.get("fit", "fill")).lower()
    grow = float(element.get("grow", 1.0) or 1.0)

    try:
        workbook_bytes = data.store.workbook_bytes(workbook)
    except FileNotFoundError as exc:
        logger.warning("North comparison skipped: %s", exc)
        print(f">>> North comparison skipped — visualizations workbook missing ({exc})")
        return False

    available = []
    try:
        available = list(data.sheet_names(workbook))
    except Exception:
        available = []

    try:
        sheet_name = resolve_sheet_name(
            workbook_bytes,
            sheet=element.get("sheet"),
            sheet_index=element.get("sheet_index"),
            sheet_match=element.get("sheet_match", ["scorecard comparison"]),
            sheet_match_index=int(element.get("sheet_match_index", 0) or 0),
            available=available or None,
        )
    except Exception as exc:
        logger.warning("Could not resolve Scorecard Comparison sheet: %s", exc)
        print(f">>> North comparison skipped — Scorecard Comparison sheet not found ({exc})")
        return False

    try:
        table = detect_scorecard_comparison_table(workbook_bytes, sheet_name)
    except Exception as exc:
        logger.warning("North comparison table detection failed: %s", exc)
        print(f">>> North comparison detection failed: {exc}")
        return False

    try:
        png = _capture_comparison_png(
            workbook_bytes, sheet_name, table, prefer_com=prefer_com
        )
    except Exception as exc:
        logger.warning("North comparison capture failed: %s", exc)
        print(f">>> North comparison capture failed: {exc}")
        return False

    left = int(element.get("left", COMPARISON_LEFT))
    top = int(element.get("top", COMPARISON_TOP))
    width = int(element.get("max_width", COMPARISON_WIDTH))
    height = int(element.get("max_height", COMPARISON_HEIGHT))

    if replace_existing:
        removed = _remove_slide_pictures(slide)
        if removed:
            logger.info("Removed %s picture(s) before North comparison image", removed)

    # Debug preview for Windows review runs.
    try:
        out_dir = Path(getattr(data.store, "base_dir", Path("."))) / "output"
        out_dir.mkdir(parents=True, exist_ok=True)
        with Image.open(io.BytesIO(png)) as img:
            debug = out_dir / f"_debug_north_comparison_{img.width}x{img.height}.png"
            img.save(debug)
            print(
                f">>> North comparison from {sheet_name}!{table.range_address} "
                f"({img.width}x{img.height}) -> {debug.name}"
            )
    except Exception:
        pass

    place_picture_on_slide(
        slide,
        png,
        left=left,
        top=top,
        max_width=width,
        max_height=height,
        fit=fit,
        grow=grow,
    )
    print(
        f"\n>>> Slide 15 North Scorecard Comparison: placed table screenshot "
        f"from {workbook}/{sheet_name}!{table.range_address}\n"
    )
    return True
