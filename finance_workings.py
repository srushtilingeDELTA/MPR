"""Finance slide (PPT 8) from New GSE MPR Workings.xlsx → FINANCE tab.

Screenshots the Regions scorecard block:
  Regions | BUDGET $000s | OVERTIME | TOTAL HOURS
  (MTD / YTD / Score headers + Actual / (B)/W Goal data rows)
"""

from __future__ import annotations

import io
import logging
import re
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from pptx.enum.shapes import MSO_SHAPE_TYPE

from scorecard_screenshots import (
    _find_sheet_name,
    _range_address,
    _validate_capture,
    capture_range_png,
    place_picture_on_slide,
    resolve_sheet_name,
)

logger = logging.getLogger(__name__)

# Template picture slot on slide 8 (from GSE MPR template).
FINANCE_TABLE_BOX = (206_375, 929_805, 11_779_250, 4_457_700)

# Each metric block is typically: Actual, (B)/W Goal, Actual, (B)/W Goal, MTD, YTD.
_METRICS_PER_BLOCK = 6
_END_TOKENS = (
    "notes",
    "leading issues",
    "action plan",
    "action plans",
    "comments",
)


def _cell_str(ws, row: int, col: int) -> str:
    val = ws.cell(row, col).value
    if val is None:
        return ""
    return str(val).strip()


def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").casefold()).strip()


def _row_blob(ws, row: int, max_col: int) -> str:
    return " ".join(_norm(_cell_str(ws, row, c)) for c in range(1, max_col + 1))


def _merged_max_col(ws, row: int, col: int) -> int:
    """If (row, col) is inside a merge, return the merge's max column."""
    try:
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                return int(merged.max_col)
    except Exception:
        pass
    return col


def _find_section_cols(ws, row: int, scan_cols: int) -> dict[str, int]:
    """Map section keys to the column where their title appears on `row`."""
    found: dict[str, int] = {}
    for col in range(1, scan_cols + 1):
        text = _norm(_cell_str(ws, row, col))
        if not text:
            continue
        if "region" in text and "regions" not in found:
            found["regions"] = col
        elif "budget" in text and "budget" not in found:
            found["budget"] = col
        elif "overtime" in text and "overtime" not in found:
            found["overtime"] = col
        elif ("total hours" in text or "total hour" in text) and "total_hours" not in found:
            found["total_hours"] = col
    # Second pass: TOTAL / HOURS split across adjacent cells.
    if "total_hours" not in found:
        for col in range(1, scan_cols):
            a = _norm(_cell_str(ws, row, col))
            b = _norm(_cell_str(ws, row, col + 1))
            if a == "total" and "hour" in b:
                found["total_hours"] = col
                break
    return found


def _content_end_col(ws, start_row: int, start_col: int, min_end: int, scan_cols: int) -> int:
    """Extend past min_end while header rows still have values (no early spacer break)."""
    end_col = min_end
    header_rows = (start_row, start_row + 1, start_row + 2, start_row + 3)
    for col in range(min_end, scan_cols + 1):
        if any(_cell_str(ws, r, col) for r in header_rows):
            end_col = col
    try:
        for merged in ws.merged_cells.ranges:
            if merged.max_row < start_row or merged.min_row > start_row + 3:
                continue
            if merged.min_col > end_col + 1:
                continue
            if merged.min_col >= start_col:
                end_col = max(end_col, int(merged.max_col))
    except Exception:
        pass
    return end_col


def _find_finance_header(ws, *, max_row: int = 50, max_col: int = 60) -> tuple[int, int, int] | None:
    """Locate Regions / BUDGET / OVERTIME / TOTAL HOURS.

    Returns (start_row, start_col, end_col) covering all three metric blocks.
    """
    # max_column can under-report when far-right cells are styled-only; always scan far enough.
    scan_cols = max(int(ws.max_column or 1), 40, min(max_col, 60))

    for row in range(1, min(int(ws.max_row or max_row), max_row) + 1):
        blob = _row_blob(ws, row, scan_cols)
        if "region" not in blob or "budget" not in blob:
            continue
        if "overtime" not in blob and "total hour" not in blob:
            continue

        sections = _find_section_cols(ws, row, scan_cols)
        if "regions" not in sections or "budget" not in sections:
            continue

        start_col = sections["regions"]
        budget_col = sections["budget"]
        overtime_col = sections.get("overtime")
        total_col = sections.get("total_hours")

        # Infer TOTAL HOURS column when the title sits on nearby header rows.
        if total_col is None:
            for r in (row, row + 1, row + 2):
                sec = _find_section_cols(ws, r, scan_cols)
                if "total_hours" in sec:
                    total_col = sec["total_hours"]
                    break

        # Block widths from title positions (Budget → OT → Total Hours).
        if overtime_col is not None and total_col is not None:
            budget_width = max(_METRICS_PER_BLOCK, overtime_col - budget_col)
            ot_width = max(_METRICS_PER_BLOCK, total_col - overtime_col)
            total_width = max(_METRICS_PER_BLOCK, ot_width, budget_width)
            min_end = max(
                _merged_max_col(ws, row, total_col),
                total_col + total_width - 1,
            )
        elif total_col is not None:
            min_end = max(
                _merged_max_col(ws, row, total_col),
                total_col + _METRICS_PER_BLOCK - 1,
            )
        elif overtime_col is not None:
            # TOTAL HOURS title missing — still reserve a full third block after OT.
            budget_width = max(_METRICS_PER_BLOCK, overtime_col - budget_col)
            min_end = overtime_col + budget_width + 1 + budget_width - 1
            logger.warning(
                "FINANCE header has OVERTIME but no TOTAL HOURS title; "
                "extending capture through col %s for the third block",
                min_end,
            )
            print(
                f">>> WARNING: TOTAL HOURS title not found — extending FINANCE capture "
                f"through column {min_end} to include the third block"
            )
        else:
            min_end = budget_col + (3 * _METRICS_PER_BLOCK) + 2

        end_col = _content_end_col(ws, row, start_col, min_end, scan_cols)
        # Hard guarantee: never end before TOTAL HOURS title + metric columns.
        if total_col is not None:
            end_col = max(end_col, total_col + _METRICS_PER_BLOCK - 1)
            end_col = max(end_col, _merged_max_col(ws, row, total_col))

        print(
            f">>> FINANCE sections: Regions=col{start_col}, BUDGET=col{budget_col}, "
            f"OVERTIME={'col' + str(overtime_col) if overtime_col else 'missing'}, "
            f"TOTAL HOURS={'col' + str(total_col) if total_col else 'missing'} "
            f"→ capture through col {end_col}"
        )
        return row, start_col, end_col
    return None


def _discover_finance_table(workbook_bytes: bytes, sheet_name: str) -> tuple[int, int, int, int]:
    """Locate the Regions BUDGET / OVERTIME / TOTAL HOURS table (headers + data)."""
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    try:
        match = _find_sheet_name(list(wb.sheetnames), sheet_name) or sheet_name
        ws = wb[match]
        hdr = _find_finance_header(ws)
        if hdr is None:
            raise ValueError(
                f"Could not find Regions/BUDGET/OVERTIME/TOTAL HOURS header on {sheet_name!r}"
            )

        start_row, start_col, end_col = hdr
        # Include MTD/YTD/Score and Actual/(B)/W Goal sub-header rows when present.
        header_rows = 1
        for extra in (1, 2):
            blob = _row_blob(ws, start_row + extra, end_col)
            if any(token in blob for token in ("mtd", "ytd", "score", "actual", "goal")):
                header_rows = extra + 1
            else:
                break

        end_row = start_row + header_rows - 1
        max_scan = min(int(ws.max_row or 80), start_row + 40)
        blank_streak = 0
        for row in range(start_row + header_rows, max_scan + 1):
            label = _cell_str(ws, row, start_col).casefold()
            row_has = any(_cell_str(ws, row, col) for col in range(start_col, end_col + 1))
            if not row_has:
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue
            blank_streak = 0
            if any(token in label for token in _END_TOKENS):
                break
            if label and label in {"finance", "kpi", "yr_nb", "entity"}:
                break
            end_row = row

        if end_row < start_row + header_rows:
            end_row = min(max_scan, start_row + header_rows + 8)

        # If any data cell exists further right (TOTAL HOURS trailing cols), include it.
        for row in range(start_row, end_row + 1):
            for col in range(end_col + 1, end_col + 10):
                if _cell_str(ws, row, col):
                    end_col = col

        addr = _range_address(start_row, end_row, start_col, end_col)
        logger.info("FINANCE Regions table %s!%s", match, addr)
        print(
            f">>> FINANCE table range: {match}!{addr} "
            f"(Regions / BUDGET / OVERTIME / TOTAL HOURS)"
        )
        return start_row, end_row, start_col, end_col
    finally:
        wb.close()


def _remove_finance_content_pictures(slide) -> int:
    """Remove the template finance overview picture (keep title / footer text)."""
    removed = 0
    left, top, width, height = FINANCE_TABLE_BOX
    band_bottom = top + height + 250_000
    for shape in list(slide.shapes):
        if shape.shape_type != MSO_SHAPE_TYPE.PICTURE:
            continue
        if int(shape.top) >= band_bottom:
            continue
        if int(shape.top) + int(shape.height) < top - 100_000:
            continue
        shape._element.getparent().remove(shape._element)
        removed += 1
    return removed


def apply_finance_workings_panels(slide, data, element: dict) -> bool:
    """Fill slide 8 from Workings!FINANCE Regions/Budget/OT/Hours table screenshot."""
    workbook = element.get("workbook", "workings")
    prefer_com = bool(element.get("prefer_excel_com", True))
    # contain keeps TOTAL HOURS visible; fill can stretch-crop wide grids.
    fit = str(element.get("fit", "contain")).lower()

    try:
        workbook_bytes = data.store.workbook_bytes(workbook)
    except FileNotFoundError as exc:
        logger.warning("FINANCE workings screenshot skipped: %s", exc)
        return False

    available = []
    try:
        available = list(data.sheet_names(workbook))
    except Exception:
        available = []

    try:
        sheet_name = resolve_sheet_name(
            workbook_bytes,
            sheet=element.get("sheet", "FINANCE"),
            sheet_index=element.get("sheet_index"),
            sheet_match=element.get("sheet_match", ["FINANCE", "Finance"]),
            sheet_match_index=int(element.get("sheet_match_index", 0) or 0),
            available=available or None,
        )
    except Exception as exc:
        logger.warning("Could not resolve Workings FINANCE sheet: %s", exc)
        return False

    try:
        start_row, end_row, start_col, end_col = _discover_finance_table(workbook_bytes, sheet_name)
        png = capture_range_png(
            workbook_bytes,
            sheet_name,
            start_row,
            end_row,
            start_col,
            end_col,
            prefer_excel_com=prefer_com,
        )
        png = _validate_capture(
            png,
            label=f"FINANCE table {sheet_name}",
            min_w=120,
            min_h=40,
            require_wide=False,
        )
    except Exception as exc:
        logger.warning("FINANCE table capture failed: %s", exc)
        print(f">>> ERROR: FINANCE table screenshot failed: {exc}")
        return False

    out_dir = Path(getattr(data.store, "base_dir", Path("."))) / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    removed = _remove_finance_content_pictures(slide)
    logger.info("Slide 8 cleared %s picture(s) before FINANCE screenshot", removed)

    left, top, width, height = FINANCE_TABLE_BOX
    place_picture_on_slide(
        slide,
        png,
        left=left,
        top=top,
        max_width=width,
        max_height=height,
        fit=fit,
    )

    try:
        with Image.open(io.BytesIO(png)) as img:
            debug = out_dir / f"_debug_finance_table_{img.width}x{img.height}.png"
            img.save(debug)
            print(
                f">>> FINANCE Regions/BUDGET/OVERTIME/TOTAL HOURS screenshot placed "
                f"from workings/{sheet_name} ({img.width}x{img.height}, fit={fit}) -> {debug.name}"
            )
    except Exception:
        print(f">>> FINANCE table screenshot placed from workings/{sheet_name}")

    print(
        f"\n>>> Slide 8 Finance: full Regions/BUDGET/OVERTIME/TOTAL HOURS screenshot "
        f"from workings/{sheet_name}\n"
    )
    return True
