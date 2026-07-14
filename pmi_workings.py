"""PMI Compliance slide (PPT 11) from New GSE MPR Workings.xlsx → PMI tab.

Only these pieces from the PMI tab:
  1. MOTORIZED / STATIONARY PMI table(s) — full width (WEIGHT…YE or Regions/MTD…Score)
  2. Motorized and Stationary Excel graphs

Everything else on the sheet (Non-Motorized, other KPIs, etc.) is ignored.
Leading Issues / Action Plans are cleared to empty editable text boxes.
"""

from __future__ import annotations

import io
import logging
import re
import tempfile
import time
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from pptx.enum.shapes import MSO_SHAPE_TYPE

from gir_panels import clear_leading_action_narrative
from people_workings import (
    _com_chart_title,
    _screenshot_excel_chart,
)
from scorecard_screenshots import (
    _find_com_worksheet,
    _find_sheet_name,
    _open_excel_workbook,
    _range_address,
    _validate_capture,
    capture_range_png,
    place_picture_on_slide,
    resolve_sheet_name,
)

logger = logging.getLogger(__name__)

# Template OLE / table slot (Object 3 on PMI COMPLIANCE slide).
PMI_TABLE_BOX = (314_628, 1_073_741, 7_994_650, 3_937_001)

# Bottom chart slots.
PMI_CHART_BOXES = {
    "motorized": (1_331_718, 5_353_707, 4_392_295, 1_353_312),
    "stationary": (5_706_868, 5_289_350, 4_389_120, 1_355_345),
}

PMI_CHART_SPECS = [
    {"key": "motorized", "title": "Motorized", "match": ["motorized", "motorised", "pm (m)"]},
    {"key": "stationary", "title": "Stationary", "match": ["stationary", "pm (s)"]},
]

_STOP_AFTER_STATIONARY = (
    "non-motor",
    "non motor",
    "nme",
    "leading issue",
    "action plan",
    "notes",
    "month",
)


def _cell_str(ws, row: int, col: int) -> str:
    val = ws.cell(row, col).value
    if val is None:
        return ""
    return str(val).strip()


def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").casefold()).strip()


def _row_blob(ws, row: int, max_col: int) -> str:
    return " ".join(_norm(_cell_str(ws, row, c)) for c in range(1, max_col + 1))


def _find_weight_kpi_header(ws, *, max_row: int = 60, max_col: int = 40) -> tuple[int, int, int] | None:
    """Find WEIGHT | KPI | TOTAL SCORE | JAN…DEC | YTD | YE header.

    Returns (header_row, start_col, end_col).
    """
    scan_cols = max(int(ws.max_column or 1), 30)
    scan_cols = min(max(scan_cols, 30), max_col)
    for row in range(1, min(int(ws.max_row or max_row), max_row) + 1):
        blob = _row_blob(ws, row, scan_cols)
        if "weight" not in blob or "kpi" not in blob:
            continue
        if "total score" not in blob and "score" not in blob:
            continue

        start_col = None
        end_col = 1
        month_cols = 0
        weight_col = None
        for col in range(1, scan_cols + 1):
            text = _norm(_cell_str(ws, row, col))
            if not text:
                continue
            if "weight" in text and weight_col is None:
                weight_col = col
            if text in {
                "jan",
                "feb",
                "mar",
                "apr",
                "may",
                "jun",
                "jul",
                "aug",
                "sep",
                "oct",
                "nov",
                "dec",
                "ytd",
                "ye",
            } or text[:3] in {
                "jan",
                "feb",
                "mar",
                "apr",
                "may",
                "jun",
                "jul",
                "aug",
                "sep",
                "oct",
                "nov",
                "dec",
            }:
                month_cols += 1
                end_col = max(end_col, col)
            if any(token in text for token in ("weight", "kpi", "total score", "score", "ytd", "ye")):
                end_col = max(end_col, col)

        if weight_col is None:
            continue
        # Include category label column to the left of WEIGHT when present.
        start_col = weight_col
        if weight_col > 1:
            start_col = weight_col - 1

        # Always span through YE / last month header present on this row.
        for col in range(end_col, scan_cols + 1):
            if _cell_str(ws, row, col):
                end_col = col
        # Guarantee room for 12 months + YTD + YE beyond WEIGHT.
        end_col = max(end_col, weight_col + 15)

        print(
            f">>> PMI WEIGHT/KPI header at row {row}, cols {start_col}-{end_col} "
            f"(month-like headers seen={month_cols})"
        )
        return row, start_col, end_col
    return None


def _section_row(ws, token: str, *, start_row: int, end_scan: int, start_col: int, end_col: int) -> int | None:
    token = token.casefold()
    for row in range(start_row, end_scan + 1):
        for col in range(start_col, end_col + 1):
            text = _norm(_cell_str(ws, row, col))
            if token in text and "non-" not in text and "non " not in text:
                return row
        # Also check far-left category column.
        for col in range(max(1, start_col - 1), start_col + 1):
            text = _norm(_cell_str(ws, row, col))
            if text == token or text.startswith(token):
                if "non-" in text or "non " in text:
                    continue
                return row
    return None


def _discover_motorized_stationary_scorecard(
    ws, sheet_name: str
) -> tuple[int, int, int, int] | None:
    """MOTORIZED + STATIONARY blocks in WEIGHT/KPI/JAN…YE layout."""
    hdr = _find_weight_kpi_header(ws)
    if hdr is None:
        return None
    header_row, start_col, end_col = hdr
    max_scan = min(int(ws.max_row or 120), header_row + 80)

    mot = _section_row(
        ws, "motorized", start_row=header_row, end_scan=max_scan, start_col=start_col, end_col=end_col
    )
    sta = _section_row(
        ws, "stationary", start_row=header_row, end_scan=max_scan, start_col=start_col, end_col=end_col
    )
    if mot is None and sta is None:
        # Banner title on/above header.
        for row in range(max(1, header_row - 3), header_row + 1):
            blob = _row_blob(ws, row, end_col)
            if "motorized" in blob and "stationary" in blob:
                mot = header_row + 1
                break
        if mot is None:
            return None

    start_row = header_row
    # Prefer starting at the earlier of motorized section / header.
    if mot is not None:
        start_row = min(start_row, mot)
    if sta is not None:
        start_row = min(start_row, sta)
    start_row = min(start_row, header_row)

    # End at last stationary-related row, before NON-MOTORIZED / other sections.
    end_row = max(mot or header_row, sta or header_row)
    blank_streak = 0
    for row in range(end_row, max_scan + 1):
        blob = _row_blob(ws, row, end_col)
        label = _norm(_cell_str(ws, row, start_col))
        left = _norm(_cell_str(ws, row, max(1, start_col)))
        if any(token in blob for token in _STOP_AFTER_STATIONARY) and row > end_row:
            # Allow "overall total score" after stationary; stop on non-motorized.
            if "non-motor" in blob or "non motor" in blob or left.startswith("nme"):
                break
            if "month" in label or label == "month":
                break
            if "leading" in blob or "action plan" in blob:
                break
        row_has = any(_cell_str(ws, row, col) for col in range(start_col, end_col + 1))
        if not row_has:
            blank_streak += 1
            if blank_streak >= 2 and row > end_row + 1:
                break
            continue
        blank_streak = 0
        # Include OVERALL TOTAL SCORE if it appears before non-motorized.
        if "overall" in blob and "score" in blob:
            end_row = row
            continue
        # Keep extending while still in motorized/stationary/plan-actual-score rows.
        if any(
            token in blob
            for token in (
                "motorized",
                "stationary",
                "plan",
                "actual",
                "percent",
                "score",
                "pmi",
            )
        ):
            end_row = row
            continue
        # Stop when a new major section begins.
        if row > end_row and any(
            token in left
            for token in ("finance", "people", "safety", "customer", "operations", "gir", "ea", "isr")
        ):
            break
        if row_has and row <= end_row + 4:
            end_row = row

    # Extend end_col across data rows (YTD/YE values).
    for row in range(start_row, end_row + 1):
        for col in range(end_col, end_col + 6):
            if _cell_str(ws, row, col):
                end_col = col

    addr = _range_address(start_row, end_row, start_col, end_col)
    print(f">>> PMI MOTORIZED/STATIONARY scorecard: {sheet_name}!{addr}")
    return start_row, end_row, start_col, end_col


def _discover_motorized_stationary_regions(
    ws, sheet_name: str
) -> tuple[int, int, int, int] | None:
    """Fallback: Regions | MTD/YTD/Score tables for Motorized/Stationary only (no Non-Motorized)."""
    scan_cols = min(max(int(ws.max_column or 20), 20), 40)
    header = None
    for row in range(1, min(int(ws.max_row or 40), 40) + 1):
        blob = _row_blob(ws, row, scan_cols)
        if "region" not in blob:
            continue
        next_blob = _row_blob(ws, row + 1, scan_cols)
        if "mtd" not in blob and "mtd" not in next_blob and "actual" not in next_blob:
            continue
        for col in range(1, scan_cols + 1):
            if "region" in _norm(_cell_str(ws, row, col)):
                header = (row, col)
                break
        if header:
            break
    if header is None:
        return None

    start_row, start_col = header
    # Only the first (Motorized/Stationary) metric block — stop before Non-Motorized columns.
    end_col = start_col
    non_motor_col = None
    for row in range(max(1, start_row - 2), start_row + 3):
        for col in range(start_col, scan_cols + 1):
            text = _norm(_cell_str(ws, row, col))
            if "non-motor" in text or "non motor" in text or text == "nme":
                non_motor_col = col if non_motor_col is None else min(non_motor_col, col)

    for row in (start_row, start_row + 1):
        for col in range(start_col, scan_cols + 1):
            if non_motor_col is not None and col >= non_motor_col:
                break
            if _cell_str(ws, row, col):
                end_col = max(end_col, col)

    # Detect a second MTD/YTD/Score block to the right (Non-Motorized without a title).
    if non_motor_col is None:
        second_block = None
        for col in range(start_col + 7, scan_cols + 1):
            top = _norm(_cell_str(ws, start_row, col))
            sub = _norm(_cell_str(ws, start_row + 1, col))
            if top in {"mtd", "ytd", "score"} or sub in {"actual", "b/w goal", "b/(w) goal", "goal"}:
                # Require a spacer/gap before this repeated block.
                gap = all(not _cell_str(ws, start_row, c) and not _cell_str(ws, start_row + 1, c) for c in range(col - 1, col))
                if gap or col >= start_col + 8:
                    second_block = col
                    break
        if second_block is not None:
            end_col = min(end_col, second_block - 1)
            # Walk left over empty spacer cols.
            while end_col > start_col and not any(
                _cell_str(ws, r, end_col) for r in (start_row, start_row + 1)
            ):
                end_col -= 1

    end_col = max(end_col, start_col + 6)
    if non_motor_col is not None:
        end_col = min(end_col, non_motor_col - 1)

    header_rows = 1
    for extra in (1, 2):
        blob = _row_blob(ws, start_row + extra, end_col)
        if any(token in blob for token in ("mtd", "ytd", "score", "actual", "goal")):
            header_rows = extra + 1
        else:
            break

    end_row = start_row + header_rows - 1
    max_scan = min(int(ws.max_row or 80), start_row + 40)
    blank_streak = 0
    for row in range(start_row + header_rows, max_scan + 1):
        label = _norm(_cell_str(ws, row, start_col))
        row_has = any(_cell_str(ws, row, col) for col in range(start_col, end_col + 1))
        if not row_has:
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0
        if any(token in label for token in _STOP_AFTER_STATIONARY):
            break
        if label in {"month", "kpi", "weight"}:
            break
        end_row = row
        # Prefer stopping after SYSTEM / STATIONARY summary rows once seen.
        if label in {"system", "stationary", "motorized"} and row > start_row + header_rows + 8:
            # keep going to include SYSTEM if present after STATIONARY
            pass

    addr = _range_address(start_row, end_row, start_col, end_col)
    print(f">>> PMI MOTORIZED/STATIONARY regions table: {sheet_name}!{addr}")
    return start_row, end_row, start_col, end_col


def _discover_pmi_table(workbook_bytes: bytes, sheet_name: str) -> tuple[int, int, int, int]:
    """Locate only MOTORIZED / STATIONARY PMI tables on the PMI sheet."""
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    try:
        match = _find_sheet_name(list(wb.sheetnames), sheet_name) or sheet_name
        ws = wb[match]

        # Prefer the monthly WEIGHT/KPI scorecard blocks for Motorized + Stationary.
        found = _discover_motorized_stationary_scorecard(ws, match)
        if found is None:
            found = _discover_motorized_stationary_regions(ws, match)
        if found is None:
            raise ValueError(
                f"Could not find MOTORIZED / STATIONARY PMI tables on {sheet_name!r}"
            )

        start_row, end_row, start_col, end_col = found
        logger.info(
            "PMI Motorized/Stationary table %s!%s",
            match,
            _range_address(start_row, end_row, start_col, end_col),
        )
        return start_row, end_row, start_col, end_col
    finally:
        wb.close()


def _match_pmi_chart(title: str) -> dict | None:
    lower = (title or "").casefold()
    # Ignore non-motorized / other charts.
    if "non-motor" in lower or "non motor" in lower or "nme" in lower:
        return None
    for spec in PMI_CHART_SPECS:
        for token in spec["match"]:
            if token in lower:
                return spec
    return None


def _iter_sheet_charts(ws) -> list[tuple[float, float, str, object]]:
    """Collect Excel charts on a sheet as (top, left, title, obj)."""
    found: list[tuple[float, float, str, object]] = []
    seen: set[str] = set()

    def _add(obj, *, source: str) -> None:
        try:
            name = str(getattr(obj, "Name", "") or "")
        except Exception:
            name = ""
        key = name.casefold() or f"{source}:{id(obj)}"
        if key in seen:
            return
        seen.add(key)
        try:
            top = float(getattr(obj, "Top", len(found) * 100))
            left = float(getattr(obj, "Left", len(found) * 100))
        except Exception:
            top, left = float(len(found) * 100), 0.0
        title = _com_chart_title(obj) or name
        found.append((top, left, title, obj))
        print(f">>> Found PMI Excel graph ({source}): {title or 'untitled'!r}")

    try:
        count = int(ws.ChartObjects().Count)
    except Exception:
        count = 0
    for idx in range(1, count + 1):
        try:
            _add(ws.ChartObjects(idx), source=f"ChartObjects[{idx}]")
        except Exception as exc:
            logger.warning("PMI ChartObjects(%s) unreadable: %s", idx, exc)

    try:
        shape_count = int(ws.Shapes.Count)
    except Exception:
        shape_count = 0
    for idx in range(1, shape_count + 1):
        try:
            shape = ws.Shapes(idx)
            if int(getattr(shape, "Type", 0) or 0) != 3:
                continue
            _add(shape, source=f"Shapes[{idx}]")
        except Exception:
            continue

    found.sort(key=lambda item: (item[0], item[1]))
    return found


def _screenshot_pmi_charts(workbook_path: Path, sheet_name: str) -> dict[str, bytes]:
    """Screenshot only Motorized + Stationary graphs from the PMI sheet."""
    excel = None
    wb = None
    by_key: dict[str, bytes] = {}
    ordered: list[bytes] = []
    try:
        excel, wb = _open_excel_workbook(workbook_path)
        ws = _find_com_worksheet(wb, sheet_name)
        ws.Activate()
        try:
            excel.ActiveWindow.Zoom = 100
        except Exception:
            pass
        time.sleep(0.3)

        charts = _iter_sheet_charts(ws)
        print(f">>> PMI sheet: {len(charts)} Excel graph(s) found")
        scored: list[tuple[float, float, str, bytes]] = []
        for idx, (top, left, title, chart_obj) in enumerate(charts, start=1):
            # Skip charts that are clearly not Mot/Stat.
            lower = (title or "").casefold()
            if any(token in lower for token in ("non-motor", "non motor", "nme", "people", "finance")):
                print(f">>> Skipping non-PMI graph {idx}: {title!r}")
                continue
            label = f"PMI graph {idx} ({title or 'untitled'})"
            try:
                try:
                    chart_obj.Activate()
                except Exception:
                    try:
                        chart_obj.Select()
                    except Exception:
                        pass
                time.sleep(0.2)
                data = _screenshot_excel_chart(chart_obj, label=label)
                scored.append((top, left, title, data))
                print(f">>> Screenshotted PMI graph {idx}: {title!r} ({len(data)} bytes)")
            except Exception as exc:
                logger.warning("%s failed: %s", label, exc)
                print(f">>> WARNING: could not screenshot PMI graph {idx}: {exc}")

        # Left-to-right then top-to-bottom for unlabeled charts.
        scored.sort(key=lambda item: (item[0], item[1]))
        for top, left, title, data in scored:
            spec = _match_pmi_chart(title)
            if spec and spec["key"] not in by_key:
                by_key[spec["key"]] = data
                print(f">>> Matched PMI screenshot to '{spec['title']}' (title {title!r})")
            else:
                ordered.append(data)

        for spec in PMI_CHART_SPECS:
            if spec["key"] in by_key:
                continue
            if not ordered:
                break
            by_key[spec["key"]] = ordered.pop(0)
            print(f">>> Assigned position-ordered PMI screenshot to '{spec['title']}'")
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
    return by_key


def _remove_pmi_content_shapes(slide) -> int:
    """Remove OLE table object, native charts, and prior content pictures (keep logo)."""
    removed = 0
    for shape in list(slide.shapes):
        drop = False
        st = shape.shape_type
        if st in (
            MSO_SHAPE_TYPE.EMBEDDED_OLE_OBJECT,
            getattr(MSO_SHAPE_TYPE, "LINKED_OLE_OBJECT", MSO_SHAPE_TYPE.EMBEDDED_OLE_OBJECT),
        ):
            drop = True
        elif getattr(shape, "has_chart", False):
            drop = True
        elif st == MSO_SHAPE_TYPE.PICTURE and 900_000 <= int(shape.top) < 6_800_000:
            drop = True
        if not drop:
            continue
        shape._element.getparent().remove(shape._element)
        removed += 1
    return removed


def apply_pmi_workings_panels(slide, data, element: dict) -> bool:
    """Fill PMI Compliance from Workings!PMI Motorized/Stationary tables + graphs only."""
    workbook = element.get("workbook", "workings")
    prefer_com = bool(element.get("prefer_excel_com", True))
    fit = str(element.get("fit", "contain")).lower()

    try:
        workbook_bytes = data.store.workbook_bytes(workbook)
    except FileNotFoundError as exc:
        logger.warning("PMI workings screenshot skipped: %s", exc)
        return False

    available = []
    try:
        available = list(data.sheet_names(workbook))
    except Exception:
        available = []

    try:
        sheet_name = resolve_sheet_name(
            workbook_bytes,
            sheet=element.get("sheet", "PMI"),
            sheet_index=element.get("sheet_index"),
            sheet_match=element.get("sheet_match", ["PMI", "PMI COMPLIANCE"]),
            sheet_match_index=int(element.get("sheet_match_index", 0) or 0),
            available=available or None,
        )
    except Exception as exc:
        logger.warning("Could not resolve Workings PMI sheet: %s", exc)
        return False

    out_dir = Path(getattr(data.store, "base_dir", Path("."))) / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    placed = 0

    # 1) MOTORIZED / STATIONARY table screenshot only.
    table_png = None
    try:
        start_row, end_row, start_col, end_col = _discover_pmi_table(workbook_bytes, sheet_name)
        table_png = capture_range_png(
            workbook_bytes,
            sheet_name,
            start_row,
            end_row,
            start_col,
            end_col,
            prefer_excel_com=prefer_com,
        )
        table_png = _validate_capture(
            table_png,
            label=f"PMI Motorized/Stationary {sheet_name}",
            min_w=200,
            min_h=80,
            require_wide=True,
        )
    except Exception as exc:
        logger.warning("PMI Motorized/Stationary table capture failed: %s", exc)
        print(f">>> ERROR: PMI MOTORIZED/STATIONARY table screenshot failed: {exc}")
        table_png = None

    # 2) Motorized + Stationary Excel graph screenshots only.
    charts_by_key: dict[str, bytes] = {}
    if prefer_com:
        try:
            with tempfile.TemporaryDirectory(prefix="mpr_pmi_") as tmp:
                path = Path(tmp) / "workings.xlsx"
                path.write_bytes(workbook_bytes)
                charts_by_key = _screenshot_pmi_charts(path, sheet_name)
        except Exception as exc:
            logger.error("PMI Excel graph screenshots unavailable: %s", exc)
            print(f">>> ERROR: could not screenshot PMI graphs from Excel: {exc}")
    else:
        print(">>> ERROR: PMI graphs require Excel COM screenshots (prefer_excel_com=true)")

    if not table_png and not charts_by_key:
        logger.warning("No PMI Motorized/Stationary screenshots from workings/%s", sheet_name)
        return False

    removed = _remove_pmi_content_shapes(slide)
    logger.info("Slide 11 cleared %s OLE/chart/picture shape(s)", removed)

    if table_png:
        left, top, width, height = PMI_TABLE_BOX
        place_picture_on_slide(
            slide,
            table_png,
            left=left,
            top=top,
            max_width=width,
            max_height=height,
            fit=fit,
        )
        placed += 1
        try:
            with Image.open(io.BytesIO(table_png)) as img:
                debug = out_dir / f"_debug_pmi_table_{img.width}x{img.height}.png"
                img.save(debug)
                print(
                    f">>> PMI MOTORIZED/STATIONARY table screenshot placed "
                    f"from workings/{sheet_name} ({img.width}x{img.height}) -> {debug.name}"
                )
        except Exception:
            print(f">>> PMI MOTORIZED/STATIONARY table screenshot placed from workings/{sheet_name}")

    for spec in PMI_CHART_SPECS:
        png = charts_by_key.get(spec["key"])
        if png is None:
            print(f">>> WARNING: missing Excel screenshot for PMI '{spec['title']}' graph")
            continue
        left, top, width, height = PMI_CHART_BOXES[spec["key"]]
        place_picture_on_slide(
            slide,
            png,
            left=left,
            top=top,
            max_width=width,
            max_height=height,
            fit=fit,
        )
        placed += 1
        try:
            with Image.open(io.BytesIO(png)) as img:
                debug = out_dir / f"_debug_pmi_{spec['key']}_{img.width}x{img.height}.png"
                img.save(debug)
                print(
                    f">>> PMI '{spec['title']}' graph screenshot placed "
                    f"({img.width}x{img.height}) -> {debug.name}"
                )
        except Exception:
            print(f">>> PMI '{spec['title']}' graph screenshot placed")

    if bool(element.get("clear_narrative", True)):
        n = clear_leading_action_narrative(slide)
        print(f">>> PMI Leading Issues / Action Plans cleared ({n} text box(es))")

    print(
        f"\n>>> Slide 11 PMI Compliance: placed {placed} screenshot(s) from workings/{sheet_name} "
        f"(MOTORIZED/STATIONARY table={'yes' if table_png else 'no'}, "
        f"graphs={len(charts_by_key)}/2; other PMI data ignored)\n"
    )
    return placed > 0
